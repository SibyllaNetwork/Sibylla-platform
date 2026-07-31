"""
Outlet Manager — Flask Backend completo
Avvio: python -m flask --app app.main run --port 8000 --debug
"""
from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime
from sqlalchemy.orm import Session

from .database import get_db, create_tables
from .models import (
    Outlet, Sala, Tavolo, Rango, Turno,
    Allergene, TipoMenu, CategoriaMenu, CategoriaCliente,
    VoceMenu, PrezzoSpeciale, VoceMenuOutlet, MenuDelGiorno,
    Stampante, VoceMenuStampante, Monitor, VoceMenuMonitor,
    WebMenu, WebMenuVoce,
    Cliente, Wallet, WalletTransazione,
    Prenotazione, Comanda, RigaComanda, voce_allergeni,
    ConfigEmail,
    ConfigMobileWallet
)

app = Flask(__name__)
CORS(app, origins=["http://localhost:3000","http://localhost:5173"], supports_credentials=True)

with app.app_context():
    create_tables()

# ── Seed dati di base ────────────────────────────────────────────────────────
def _seed():
    db = get_db()
    try:
        # Allergeni standard EU (14)
        if db.query(Allergene).count() == 0:
            for cod, nome, desc in [
                ("A","Glutine","Cereali contenenti glutine: grano, segale, orzo, avena..."),
                ("B","Crostacei","Granchio, gambero, aragosta, scampi..."),
                ("C","Uova","Uova e prodotti a base di uova"),
                ("D","Pesce","Pesce e prodotti a base di pesce"),
                ("E","Arachidi","Arachidi e prodotti a base di arachidi"),
                ("F","Soia","Soia e prodotti a base di soia"),
                ("G","Latte","Latte e prodotti a base di latte (lattosio)"),
                ("H","Frutta a guscio","Mandorle, nocciole, noci, anacardi, pistacchi..."),
                ("I","Sedano","Sedano e prodotti a base di sedano"),
                ("J","Senape","Senape e prodotti a base di senape"),
                ("K","Semi di sesamo","Semi di sesamo e prodotti a base di sesamo"),
                ("L","Anidride solforosa","Solfiti a concentrazioni > 10 mg/kg o 10 mg/l"),
                ("M","Lupini","Lupini e prodotti a base di lupini"),
                ("N","Molluschi","Molluschi e prodotti a base di molluschi"),
            ]:
                db.add(Allergene(codice=cod, nome=nome, descrizione=desc))
            db.commit()

        # Tipi menu
        if db.query(TipoMenu).count() == 0:
            for i, (nome, colore) in enumerate([
                ("Food","#dc2626"), ("Beverage","#7c3aed"), ("Cantina","#9f1239")
            ]):
                db.add(TipoMenu(nome=nome, colore=colore, ordine=i))
            db.commit()

        print("✅ Dati base inizializzati")
    except Exception as e:
        db.rollback(); print(f"⚠️  Seed: {e}")
    finally:
        db.close()

with app.app_context():
    _seed()


def ok(data, status=200): return jsonify(data), status
def err(msg, status=400): return jsonify({"error": msg}), status
def body(): return request.get_json(force=True, silent=True) or {}


# ═══════════════════════════════════════════════════════════════════════════
# OUTLETS
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/outlets")
def get_outlets():
    db = get_db()
    try:
        return ok([o.to_dict() for o in db.query(Outlet).order_by(Outlet.nome).all()])
    finally: db.close()

@app.post("/api/outlets")
def create_outlet():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        o = Outlet(nome=d["nome"], tipo=d.get("tipo","ristorante"),
                   struttura=d.get("struttura"), indirizzo=d.get("indirizzo"),
                   telefono=d.get("telefono"), email=d.get("email"))
        db.add(o); db.commit(); db.refresh(o)
        return ok(o.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.get("/api/outlets/<int:oid>")
def get_outlet(oid):
    db = get_db()
    try:
        o = db.get(Outlet, oid)
        return ok(o.to_dict()) if o else err("Non trovato", 404)
    finally: db.close()

@app.put("/api/outlets/<int:oid>")
def update_outlet(oid):
    db = get_db()
    try:
        o = db.get(Outlet, oid)
        if not o: return err("Non trovato", 404)
        d = body()
        for k in ("nome","tipo","struttura","indirizzo","telefono","email","attivo"):
            if k in d: setattr(o, k, d[k])
        db.commit(); db.refresh(o)
        return ok(o.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/outlets/<int:oid>")
def delete_outlet(oid):
    db = get_db()
    try:
        o = db.get(Outlet, oid)
        if not o: return err("Non trovato", 404)
        db.delete(o); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# SALE
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/outlets/<int:oid>/sale")
def get_sale(oid):
    db = get_db()
    try:
        return ok([s.to_dict() for s in db.query(Sala).filter(Sala.outlet_id==oid).all()])
    finally: db.close()

@app.post("/api/outlets/<int:oid>/sale")
def create_sala(oid):
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        s = Sala(outlet_id=oid, nome=d["nome"],
                 num_tavoli=d.get("num_tavoli",0), capienza_max=d.get("capienza_max",0))
        db.add(s); db.flush()
        # Crea tavoli automaticamente se specificato
        n = int(d.get("num_tavoli", 0))
        colors = ["#ec4899","#22c55e","#f97316","#3b82f6","#eab308","#a855f7",
                  "#06b6d4","#f43f5e","#84cc16","#fb923c","#818cf8","#14b8a6"]
        for i in range(1, n+1):
            db.add(Tavolo(sala_id=s.id, numero=f"{i:03d}", capienza=d.get("capienza_tavolo",4),
                          hat_color=colors[(i-1)%len(colors)]))
        db.commit(); db.refresh(s)
        return ok(s.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/sale/<int:sid>")
def update_sala(sid):
    db = get_db()
    try:
        s = db.get(Sala, sid)
        if not s: return err("Non trovato", 404)
        d = body()
        for k in ("nome","num_tavoli","capienza_max","attiva"):
            if k in d: setattr(s, k, d[k])
        db.commit(); db.refresh(s)
        return ok(s.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/sale/<int:sid>")
def delete_sala(sid):
    db = get_db()
    try:
        s = db.get(Sala, sid)
        if not s: return err("Non trovato", 404)
        db.delete(s); db.commit()
        return ok({"detail": "eliminata"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# TAVOLI
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/sale/<int:sid>/tavoli")
def get_tavoli(sid):
    db = get_db()
    try:
        tavoli = db.query(Tavolo).filter(Tavolo.sala_id==sid).order_by(Tavolo.numero).all()
        # Auto-clean stale tavolo_unito_id: clear if table is disponibile or
        # if the partner no longer points back (broken/orphan link)
        dirty = False
        tav_by_id = {t.id: t for t in tavoli}
        for t in tavoli:
            if t.tavolo_unito_id:
                partner = tav_by_id.get(t.tavolo_unito_id)
                stale = (
                    t.status == "disponibile" or           # free table shouldn't be linked
                    partner is None or                      # partner doesn't exist
                    (partner.status == "disponibile" and    # both free = stale
                     t.status == "disponibile")
                )
                if stale:
                    t.tavolo_unito_id = None
                    dirty = True
        if dirty:
            db.commit()
        return ok([t.to_dict() for t in tavoli])
    finally: db.close()

@app.post("/api/sale/<int:sid>/tavoli")
def create_tavolo(sid):
    db = get_db()
    try:
        d = body()
        if not d.get("numero"): return err("numero obbligatorio")
        t = Tavolo(sala_id=sid, numero=d["numero"], capienza=d.get("capienza",4),
                   hat_color=d.get("hat_color","#94a3b8"))
        db.add(t); db.commit(); db.refresh(t)
        return ok(t.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/tavoli/<int:tid>")
def update_tavolo_full(tid):
    db = get_db()
    try:
        t = db.get(Tavolo, tid)
        if not t: return err("Non trovato", 404)
        d = body()
        for k in ("numero","capienza","status","coperti_attuali","rango_id",
                  "tavolo_unito_id","hat_color","bloccato"):
            if k in d: setattr(t, k, d[k])
        db.commit(); db.refresh(t)
        return ok(t.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.patch("/api/tavoli/<int:tid>")
def patch_tavolo(tid):
    db = get_db()
    try:
        t = db.get(Tavolo, tid)
        if not t: return err("Non trovato", 404)
        d = body()
        for k, v in d.items():
            if hasattr(t, k): setattr(t, k, v)
        # When table becomes disponibile, always clear union link
        if d.get("status") == "disponibile":
            t.tavolo_unito_id = None
            # Also clear partner's back-link if still pointing here
            if t.tavolo_unito_id:  # original before clear
                pass
            # Find any partner that points to this table and clear their link too
            db.query(Tavolo).filter(
                Tavolo.tavolo_unito_id == tid,
                Tavolo.status == "disponibile"
            ).update({"tavolo_unito_id": None})
        db.commit(); db.refresh(t)
        return ok(t.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/tavoli/<int:tid>")
def delete_tavolo(tid):
    db = get_db()
    try:
        t = db.get(Tavolo, tid)
        if not t: return err("Non trovato", 404)
        db.delete(t); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.post("/api/tavoli/<int:t1>/unisci/<int:t2>")
def unisci_tavoli(t1, t2):
    db = get_db()
    try:
        ta = db.get(Tavolo, t1)
        if not ta: return err("Non trovato", 404)
        ta.tavolo_unito_id = t2; db.commit()
        return ok({"detail": f"Tavolo {t1} unito a {t2}"})
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# TURNI
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/turni")
def get_turni():
    db = get_db()
    try:
        q = db.query(Turno)
        if request.args.get("outlet_id"): q = q.filter(Turno.outlet_id==int(request.args["outlet_id"]))
        if request.args.get("sala_id"):   q = q.filter(Turno.sala_id==int(request.args["sala_id"]))
        return ok([t.to_dict() for t in q.order_by(Turno.servizio, Turno.ora_inizio).all()])
    finally: db.close()

@app.post("/api/turni")
def create_turno():
    db = get_db()
    try:
        d = body()
        if not d.get("outlet_id"): return err("outlet_id obbligatorio")
        t = Turno(outlet_id=d["outlet_id"], sala_id=d.get("sala_id"),
                  servizio=d.get("servizio","Pranzo"), nome=d.get("nome","Turno 1"),
                  ora_inizio=d.get("ora_inizio","12:00"), ora_fine=d.get("ora_fine","15:00"),
                  copertura_max=d.get("copertura_max",0))
        db.add(t); db.commit(); db.refresh(t)
        return ok(t.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/turni/<int:tid>")
def update_turno(tid):
    db = get_db()
    try:
        t = db.get(Turno, tid)
        if not t: return err("Non trovato", 404)
        d = body()
        for k in ("sala_id","servizio","nome","ora_inizio","ora_fine","copertura_max","attivo"):
            if k in d: setattr(t, k, d[k])
        db.commit(); db.refresh(t)
        return ok(t.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/turni/<int:tid>")
def delete_turno(tid):
    db = get_db()
    try:
        t = db.get(Turno, tid)
        if not t: return err("Non trovato", 404)
        db.delete(t); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# ALLERGENI
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/allergeni")
def get_allergeni():
    db = get_db()
    try:
        return ok([a.to_dict() for a in db.query(Allergene).order_by(Allergene.codice).all()])
    finally: db.close()

@app.post("/api/allergeni")
def create_allergene():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        a = Allergene(codice=d.get("codice",""), nome=d["nome"], descrizione=d.get("descrizione",""))
        db.add(a); db.commit(); db.refresh(a)
        return ok(a.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/allergeni/<int:aid>")
def update_allergene(aid):
    db = get_db()
    try:
        a = db.get(Allergene, aid)
        if not a: return err("Non trovato", 404)
        d = body()
        for k in ("codice","nome","descrizione","attivo"):
            if k in d: setattr(a, k, d[k])
        db.commit(); db.refresh(a)
        return ok(a.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/allergeni/<int:aid>")
def delete_allergene(aid):
    db = get_db()
    try:
        a = db.get(Allergene, aid)
        if not a: return err("Non trovato", 404)
        db.delete(a); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# TIPI MENU
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/tipi-menu")
def get_tipi():
    db = get_db()
    try:
        return ok([t.to_dict() for t in db.query(TipoMenu).order_by(TipoMenu.ordine).all()])
    finally: db.close()

@app.post("/api/tipi-menu")
def create_tipo():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        t = TipoMenu(nome=d["nome"], colore=d.get("colore","#64748b"), ordine=d.get("ordine",0))
        db.add(t); db.commit(); db.refresh(t)
        return ok(t.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/tipi-menu/<int:tid>")
def update_tipo(tid):
    db = get_db()
    try:
        t = db.get(TipoMenu, tid)
        if not t: return err("Non trovato", 404)
        d = body()
        for k in ("nome","colore","ordine"):
            if k in d: setattr(t, k, d[k])
        db.commit(); db.refresh(t)
        return ok(t.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/tipi-menu/<int:tid>")
def delete_tipo(tid):
    db = get_db()
    try:
        t = db.get(TipoMenu, tid)
        if not t: return err("Non trovato", 404)
        db.delete(t); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# CATEGORIE MENU
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/categorie-menu")
def get_categorie():
    db = get_db()
    try:
        cats = db.query(CategoriaMenu).order_by(CategoriaMenu.ordine).all()
        inc = request.args.get("include_voci") == "1"
        return ok([c.to_dict(include_voci=inc) for c in cats])
    finally: db.close()

@app.post("/api/categorie-menu")
def create_categoria():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        c = CategoriaMenu(nome=d["nome"], tipo_id=d.get("tipo_id"),
                          colore=d.get("colore","#3b82f6"), emoji=d.get("emoji","🍽️"),
                          ordine=d.get("ordine",0))
        db.add(c); db.commit(); db.refresh(c)
        return ok(c.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/categorie-menu/<int:cid>")
def update_categoria(cid):
    db = get_db()
    try:
        c = db.get(CategoriaMenu, cid)
        if not c: return err("Non trovato", 404)
        d = body()
        for k in ("nome","tipo_id","colore","emoji","ordine"):
            if k in d: setattr(c, k, d[k])
        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/categorie-menu/<int:cid>")
def delete_categoria(cid):
    db = get_db()
    try:
        c = db.get(CategoriaMenu, cid)
        if not c: return err("Non trovato", 404)
        db.delete(c); db.commit()
        return ok({"detail": "eliminata"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# CATEGORIE CLIENTE
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/categorie-cliente")
def get_cat_cliente():
    db = get_db()
    try:
        return ok([c.to_dict() for c in db.query(CategoriaCliente).order_by(CategoriaCliente.nome).all()])
    finally: db.close()

@app.post("/api/categorie-cliente")
def create_cat_cliente():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        c = CategoriaCliente(nome=d["nome"], descrizione=d.get("descrizione"),
                             sconto_perc=d.get("sconto_perc",0.0))
        db.add(c); db.commit(); db.refresh(c)
        return ok(c.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/categorie-cliente/<int:cid>")
def update_cat_cliente(cid):
    db = get_db()
    try:
        c = db.get(CategoriaCliente, cid)
        if not c: return err("Non trovato", 404)
        d = body()
        for k in ("nome","descrizione","sconto_perc"):
            if k in d: setattr(c, k, d[k])
        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/categorie-cliente/<int:cid>")
def delete_cat_cliente(cid):
    db = get_db()
    try:
        c = db.get(CategoriaCliente, cid)
        if not c: return err("Non trovato", 404)
        db.delete(c); db.commit()
        return ok({"detail": "eliminata"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# VOCI MENU
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/voci-menu")
def get_voci():
    db = get_db()
    try:
        from sqlalchemy import or_, exists as sa_exists
        q = db.query(VoceMenu)
        if request.args.get("categoria_id"): q = q.filter(VoceMenu.categoria_id==int(request.args["categoria_id"]))
        if request.args.get("solo_attivi","1") == "1": q = q.filter(VoceMenu.attivo==True)
        if request.args.get("outlet_id"):
            oid = int(request.args["outlet_id"])
            no_outlet_set = ~sa_exists().where(VoceMenuOutlet.voce_id == VoceMenu.id)
            has_outlet    =  sa_exists().where(
                (VoceMenuOutlet.voce_id == VoceMenu.id) & (VoceMenuOutlet.outlet_id == oid)
            )
            q = q.filter(or_(no_outlet_set, has_outlet))
        if request.args.get("outlet_id"):
            from sqlalchemy import or_, exists as sa_ex
            oid = int(request.args["outlet_id"])
            q = q.filter(or_(
                ~sa_ex().where(VoceMenuOutlet.voce_id == VoceMenu.id),
                sa_ex().where((VoceMenuOutlet.voce_id == VoceMenu.id) & (VoceMenuOutlet.outlet_id == oid))
            ))
        return ok([v.to_dict() for v in q.order_by(VoceMenu.nome_it).all()])
    finally: db.close()

@app.get("/api/voci-menu/<int:vid>")
def get_voce(vid):
    db = get_db()
    try:
        v = db.get(VoceMenu, vid)
        return ok(v.to_dict()) if v else err("Non trovato", 404)
    finally: db.close()


def _save_voce_outlets(db, voce_id, outlet_ids):
    """Replace outlet links for a voce."""
    db.query(VoceMenuOutlet).filter(VoceMenuOutlet.voce_id == voce_id).delete()
    for oid in (outlet_ids or []):
        db.add(VoceMenuOutlet(voce_id=voce_id, outlet_id=int(oid)))

@app.post("/api/voci-menu")
def create_voce():
    db = get_db()
    try:
        d = body()
        if not d.get("nome_it"): return err("nome_it obbligatorio")
        if not d.get("categoria_id"): return err("categoria_id obbligatorio")
        v = VoceMenu(
            categoria_id=d["categoria_id"], nome_it=d["nome_it"],
            nome_en=d.get("nome_en"), nome_de=d.get("nome_de"), nome_fr=d.get("nome_fr"),
            descrizione=d.get("descrizione"), prezzo=float(d.get("prezzo",0)),
        )
        # Allergeni
        for aid in d.get("allergeni_ids", []):
            a = db.get(Allergene, aid)
            if a: v.allergeni.append(a)
        db.add(v); db.flush()
        # Outlet links
        _save_voce_outlets(db, v.id, d.get("outlet_ids", []))
        # Prezzi speciali (con outlet_id opzionale)
        for ps in d.get("prezzi_speciali", []):
            if ps.get("categoria_cliente_id") or ps.get("outlet_id"):
                db.add(PrezzoSpeciale(
                    voce_id=v.id,
                    categoria_cliente_id=ps.get("categoria_cliente_id") or None,
                    outlet_id=ps.get("outlet_id") or None,
                    prezzo_override=float(ps["prezzo_override"])))
        db.commit(); db.refresh(v)
        return ok(v.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/voci-menu/<int:vid>")
def update_voce(vid):
    db = get_db()
    try:
        v = db.get(VoceMenu, vid)
        if not v: return err("Non trovato", 404)
        d = body()
        for k in ("categoria_id","nome_it","nome_en","nome_de","nome_fr","descrizione","prezzo","attivo"):
            if k in d: setattr(v, k, d[k])
        # Aggiorna allergeni
        if "allergeni_ids" in d:
            v.allergeni.clear()
            for aid in d["allergeni_ids"]:
                a = db.get(Allergene, aid)
                if a: v.allergeni.append(a)
        # Aggiorna outlet links
        if "outlet_ids" in d:
            _save_voce_outlets(db, v.id, d["outlet_ids"])
        # Aggiorna prezzi speciali (con outlet_id opzionale)
        if "prezzi_speciali" in d:
            for ps in v.prezzi_spec: db.delete(ps)
            db.flush()
            for ps in d["prezzi_speciali"]:
                if ps.get("categoria_cliente_id") or ps.get("outlet_id"):
                    db.add(PrezzoSpeciale(
                        voce_id=v.id,
                        categoria_cliente_id=ps.get("categoria_cliente_id") or None,
                        outlet_id=ps.get("outlet_id") or None,
                        prezzo_override=float(ps["prezzo_override"])))
        if "nel_web_menu" in d: v.nel_web_menu = bool(d["nel_web_menu"])
        db.commit(); db.refresh(v)
        return ok(v.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.patch("/api/categorie-menu/reorder")
def reorder_categorie():
    """Receive [{id, ordine}, ...] and update each category's sort order."""
    db = get_db()
    try:
        items = body()  # list of {id, ordine}
        for item in items:
            cat = db.get(CategoriaMenu, item["id"])
            if cat: cat.ordine = item["ordine"]
        db.commit()
        return ok({"updated": len(items)})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.patch("/api/voci-menu/reorder")
def reorder_voci():
    """Receive [{id, ordine}, ...] and update each voce's sort order."""
    db = get_db()
    try:
        items = body()  # list of {id, ordine}
        for item in items:
            v = db.get(VoceMenu, item["id"])
            if v: v.ordine = item["ordine"]
        db.commit()
        return ok({"updated": len(items)})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/voci-menu/<int:vid>")
def delete_voce(vid):
    db = get_db()
    try:
        v = db.get(VoceMenu, vid)
        if not v: return err("Non trovato", 404)
        db.delete(v); db.commit()
        return ok({"detail": "eliminata"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# MENU DEL GIORNO
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/menu-del-giorno")
def get_menu():
    db = get_db()
    try:
        q = db.query(MenuDelGiorno)
        if request.args.get("outlet_id"): q = q.filter(MenuDelGiorno.outlet_id==int(request.args["outlet_id"]))
        if request.args.get("data"):      q = q.filter(MenuDelGiorno.data==request.args["data"])
        return ok([m.to_dict() for m in q.order_by(MenuDelGiorno.data.desc()).all()])
    finally: db.close()

@app.post("/api/menu-del-giorno")
def create_menu():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        m = MenuDelGiorno(
            outlet_id=d.get("outlet_id"), nome=d["nome"],
            data=d.get("data"), prezzo_fisso=d.get("prezzo_fisso"),
            note=d.get("note")
        )
        for vid in d.get("voci_ids", []):
            v = db.get(VoceMenu, vid)
            if v: m.voci.append(v)
        db.add(m); db.commit(); db.refresh(m)
        return ok(m.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/menu-del-giorno/<int:mid>")
def update_menu(mid):
    db = get_db()
    try:
        m = db.get(MenuDelGiorno, mid)
        if not m: return err("Non trovato", 404)
        d = body()
        for k in ("nome","data","prezzo_fisso","note","attivo"):
            if k in d: setattr(m, k, d[k])
        if "voci_ids" in d:
            m.voci.clear()
            for vid in d["voci_ids"]:
                v = db.get(VoceMenu, vid)
                if v: m.voci.append(v)
        db.commit(); db.refresh(m)
        return ok(m.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/menu-del-giorno/<int:mid>")
def delete_menu(mid):
    db = get_db()
    try:
        m = db.get(MenuDelGiorno, mid)
        if not m: return err("Non trovato", 404)
        db.delete(m); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# PRENOTAZIONI
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/prenotazioni")
def get_prenotazioni():
    db = get_db()
    try:
        q = db.query(Prenotazione)
        for k in ("outlet_id","sala_id","turno_id"):
            if request.args.get(k): q = q.filter(getattr(Prenotazione,k)==int(request.args[k]))
        if request.args.get("data"): q = q.filter(Prenotazione.data==request.args["data"])
        return ok([p.to_dict() for p in q.order_by(Prenotazione.data,Prenotazione.orario).all()])
    finally: db.close()

@app.post("/api/prenotazioni")
def create_prenotazione():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        p = Prenotazione(**{k:v for k,v in d.items() if hasattr(Prenotazione,k) and k not in ("id","created_at")})
        db.add(p); db.commit(); db.refresh(p)
        if p.tavolo_id:
            t = db.get(Tavolo, p.tavolo_id)
            if t: t.status = "riservato"; db.commit()
        return ok(p.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.put("/api/prenotazioni/<int:pid>")
def update_prenotazione(pid):
    db = get_db()
    try:
        p = db.get(Prenotazione, pid)
        if not p: return err("Non trovato", 404)
        old_tavolo_id = p.tavolo_id
        d = body()
        for k in ("sala_id","tavolo_id","turno_id","data","orario","servizio",
                  "nome","telefono","email","coperti","note","confermata","is_vip"):
            if k in d: setattr(p, k, d[k])
        # If tavolo changed: free old, reserve new
        if "tavolo_id" in d:
            new_tid = d["tavolo_id"]
            if old_tavolo_id and old_tavolo_id != new_tid:
                old_t = db.get(Tavolo, old_tavolo_id)
                if old_t and old_t.status == "riservato":
                    old_t.status = "disponibile"
                    old_t.tavolo_unito_id = None
            if new_tid:
                new_t = db.get(Tavolo, int(new_tid))
                if new_t: new_t.status = "riservato"
        db.commit(); db.refresh(p)
        return ok(p.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/prenotazioni/<int:pid>")
def delete_prenotazione(pid):
    db = get_db()
    try:
        p = db.get(Prenotazione, pid)
        # Free the tavolo before deleting
        if p and p.tavolo_id:
            t = db.get(Tavolo, p.tavolo_id)
            if t and t.status == "riservato":
                t.status = "disponibile"
                t.tavolo_unito_id = None
        if not p: return err("Non trovato", 404)
        db.delete(p); db.commit()
        return ok({"detail": "eliminata"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ═══════════════════════════════════════════════════════════════════════════
# COMANDE
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/comande")
def get_comande():
    db = get_db()
    try:
        q = db.query(Comanda)
        if request.args.get("tavolo_id"): q = q.filter(Comanda.tavolo_id==int(request.args["tavolo_id"]))
        if request.args.get("outlet_id"): q = q.filter(Comanda.outlet_id==int(request.args["outlet_id"]))
        if request.args.get("status"):    q = q.filter(Comanda.status==request.args["status"])
        if request.args.get("data"):      q = q.filter(Comanda.created_at >= request.args["data"])
        return ok([c.to_dict() for c in q.order_by(Comanda.created_at.desc()).all()])
    finally: db.close()

@app.post("/api/comande")
def create_comanda():
    db = get_db()
    try:
        d = body()
        # tavolo_id is optional (outlets without tables)
        tavolo_id = d.get("tavolo_id") or None
        outlet_id = d.get("outlet_id") or None
        cat_cid   = d.get("cat_cliente_id") or None

        # Derive outlet for numero progressivo
        from sqlalchemy import func as sqlfunc
        outlet_id_for_num = outlet_id
        if not outlet_id_for_num and tavolo_id:
            t = db.get(Tavolo, int(tavolo_id))
            if t:
                s = db.get(Sala, t.sala_id)
                if s: outlet_id_for_num = s.outlet_id

        # Progressive number with outlet prefix
        if outlet_id_for_num:
            oid_int = int(outlet_id_for_num)
            o = db.get(Outlet, oid_int)
            if o:
                letters = [c for c in o.nome if c.isalpha()]
                pfx = "".join(letters[:2]).upper() if len(letters)>=2 else f"O{oid_int}"
            else:
                pfx = f"O{oid_int}"
            count = db.query(sqlfunc.count(Comanda.id)).filter(
                Comanda.outlet_id == oid_int
            ).scalar() or 0
            auto_num = f"{pfx}{count+1:03d}"
        else:
            count = db.query(sqlfunc.count(Comanda.id)).scalar() or 0
            auto_num = f"{count+1:03d}"
        numero = d.get("numero") or auto_num

        # Create comanda
        c = Comanda(
            tavolo_id      = int(tavolo_id) if tavolo_id else None,
            turno_id       = d.get("turno_id"),
            outlet_id      = int(outlet_id) if outlet_id else None,
            cat_cliente_id = int(cat_cid) if cat_cid else None,
            numero         = numero,
            coperti        = d.get("coperti", 0),
            note           = d.get("note", ""),
        )
        db.add(c); db.flush()

        # If tavolo provided, mark it as attesa_ordine
        if tavolo_id:
            t = db.get(Tavolo, int(tavolo_id))
            if t:
                t.status = "attesa_ordine"
                t.turno_occupato_id = d.get("turno_id")

        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ── Sposta comanda su altro tavolo ────────────────────────────────────────────
@app.post("/api/comande/<int:cid>/sposta")
def sposta_comanda(cid):
    """Sposta l'intera comanda su un altro tavolo (altra sala/outlet/turno)."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Comanda non trovata", 404)
        d = body()
        nuovo_tavolo_id = d.get("nuovo_tavolo_id")
        if not nuovo_tavolo_id: return err("nuovo_tavolo_id obbligatorio")

        nuovo_t = db.get(Tavolo, int(nuovo_tavolo_id))
        if not nuovo_t: return err("Tavolo destinazione non trovato", 404)
        if nuovo_t.status not in ("disponibile",):
            return err(f"Il tavolo {nuovo_t.numero} non è disponibile (status: {nuovo_t.status})")

        vecchio_tavolo_id = c.tavolo_id

        # Aggiorna comanda con nuovo tavolo (senza refresh della relazione)
        db.execute(
            __import__("sqlalchemy").text("UPDATE comande SET tavolo_id=:tid WHERE id=:cid"),
            {"tid": int(nuovo_tavolo_id), "cid": c.id}
        )

        # Libera tavolo origine
        vecchio_t = db.get(Tavolo, vecchio_tavolo_id)
        if vecchio_t:
            vecchio_t.status = "disponibile"
            vecchio_t.coperti_attuali = 0
            vecchio_t.turno_occupato_id = None
            vecchio_t.pagato = False

        # Occupa tavolo destinazione
        nuovo_t.status = "occupato"
        nuovo_t.coperti_attuali = c.coperti
        nuovo_t.turno_occupato_id = c.turno_id

        db.commit()
        # Re-fetch comanda fresh from DB to avoid duplicate righe in ORM cache
        db.expire_all()
        c_fresh = db.get(Comanda, c.id)
        return ok({"comanda": c_fresh.to_dict(), "tavolo_nuovo": nuovo_t.to_dict()})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ── Unisci tavolo con altro tavolo ────────────────────────────────────────────
@app.post("/api/comande/<int:cid>/unisci")
def unisci_comanda(cid):
    """Unisce le righe di un'altra comanda in questa e libera il tavolo fonte."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Comanda non trovata", 404)
        d = body()
        altro_tavolo_id = d.get("altro_tavolo_id")
        if not altro_tavolo_id: return err("altro_tavolo_id obbligatorio")

        t_src = db.get(Tavolo, int(altro_tavolo_id))
        if not t_src: return err("Tavolo non trovato", 404)

        t_dest = db.get(Tavolo, c.tavolo_id)

        # Cerca comanda aperta sull'altro tavolo (se esiste)
        altra = db.query(Comanda).filter(
            Comanda.tavolo_id == int(altro_tavolo_id),
            Comanda.status == "aperta"
        ).first()

        if altra and altra.id != c.id:
            # Sposta righe dell'altra comanda in questa
            for r in altra.righe:
                r.comanda_id = c.id
            altra.righe = []
            db.flush()
            # Segna altra come unita
            altra.status = "unita"
            altra.closed_at = datetime.utcnow()

        # Ricalcola totale comanda principale
        db.flush()
        c.totale = sum(r.prezzo_snapshot * r.quantita for r in c.righe)

        # Unisci i tavoli fisicamente
        # t_dest = table WITH the comanda (origin/active)
        # t_src  = other table being joined (becomes grayed/inactive)
        if t_dest:
            t_dest.tavolo_unito_id = t_src.id   # origin shows link to joined
            t_dest.coperti_attuali = (t_dest.coperti_attuali or 0) + (t_src.coperti_attuali or 0)
            t_dest.status = "occupato"           # origin stays active/occupato
        # The joined/secondary table is grayed out - points back to origin
        t_src.tavolo_unito_id = t_dest.id if t_dest else None
        t_src.status = "attesa_ordine"           # grayed but visible
        t_src.turno_occupato_id = c.turno_id

        db.commit()
        db.expire_all()
        c_fresh = db.get(Comanda, c.id)
        return ok({
            "comanda": c_fresh.to_dict(),
            "tavolo_dest": t_dest.to_dict() if t_dest else None,
            "tavolo_src": t_src.to_dict()
        })
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ── Estratto conto (dati JSON per PDF lato client) ────────────────────────────
@app.patch("/api/comande/<int:cid>")
def update_comanda(cid):
    """Aggiorna righe, note, inviato_monitor, turno_corrente di una comanda."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        d = body()
        if "note" in d: c.note = d["note"]
        if "coperti" in d: c.coperti = int(d["coperti"])
        if "inviato_monitor" in d: c.inviato_monitor = bool(d["inviato_monitor"])
        if "turno_corrente" in d: c.turno_corrente = int(d["turno_corrente"])
        if "righe" in d:
            for r in c.righe: db.delete(r)
            db.flush()
            totale = 0.0
            for item in (d["righe"] or []):
                rg = RigaComanda(
                    comanda_id      = cid,
                    voce_id         = item.get("voce_id"),
                    nome_snapshot   = item.get("nome_snapshot",""),
                    prezzo_snapshot = float(item.get("prezzo_snapshot",0)),
                    quantita        = int(item.get("quantita",1)),
                    note            = item.get("note"),
                    turno_idx       = int(item.get("turno_idx",0)),
                )
                db.add(rg)
                totale += rg.prezzo_snapshot * rg.quantita
            c.totale = totale
        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()



@app.patch("/api/comande/<int:cid>/avanza-turno")
def avanza_turno(cid):
    """Increment turno_corrente — moves the bell to the next turno on the KDS."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        c.turno_corrente = (c.turno_corrente or 0) + 1
        db.commit(); db.refresh(c)
        return ok({"turno_corrente": c.turno_corrente})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/comande/<int:cid>")
def delete_comanda(cid):
    """Annulla e cancella una comanda."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        # If tavolo linked, mark disponibile
        if c.tavolo_id:
            t = db.get(Tavolo, c.tavolo_id)
            if t:
                t.status = "disponibile"
                t.coperti_attuali = 0
                t.turno_occupato_id = None
        db.delete(c)
        db.commit()
        return ok({"deleted": cid})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.post("/api/comande/<int:cid>/riapri")
def riapri_comanda(cid):
    """Riapre una comanda chiusa."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        c.status = "aperta"
        c.tipo_chiusura = None
        c.closed_at = None
        # Re-mark tavolo as occupato if still linked
        if c.tavolo_id:
            t = db.get(Tavolo, c.tavolo_id)
            if t and t.status == "disponibile":
                t.status = "attesa_ordine"
        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.post("/api/comande/<int:cid>/chiudi")
def chiudi_comanda(cid):
    """Chiudi comanda con tipo di chiusura (alias di paga)."""
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        tipo = request.args.get("tipo_chiusura") or body().get("tipo","scontrino")
        c.status = "chiusa"
        c.tipo_chiusura = tipo
        from datetime import datetime as _dt
        c.closed_at = _dt.utcnow()
        if c.tavolo_id:
            t = db.get(Tavolo, c.tavolo_id)
            if t:
                t.status = "disponibile"
                t.coperti_attuali = 0
                t.turno_occupato_id = None
                t.tavolo_unito_id = None
        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.get("/api/comande/<int:cid>")
def get_comanda(cid):
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        return ok(c.to_dict())
    finally:
        db.close()


@app.get("/api/comande/<int:cid>/estratto-conto")
def estratto_conto(cid):
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        t = db.get(Tavolo, c.tavolo_id)
        from .models import Turno as TurnoM, Sala as SalaM, Outlet as OutletM
        turno  = db.get(TurnoM,  c.turno_id)  if c.turno_id else None
        sala   = db.get(SalaM,   t.sala_id)   if t else None
        outlet = db.get(OutletM, sala.outlet_id) if sala else None
        return ok({
            "comanda_numero": c.numero,
            "tavolo": t.numero if t else "?",
            "sala": sala.nome if sala else "",
            "outlet": outlet.nome if outlet else "",
            "turno": turno.nome if turno else "",
            "coperti": c.coperti,
            "note": c.note or "",
            "data": str(c.created_at)[:16].replace("T"," "),
            "righe": [{"nome":r.nome_snapshot,"qty":r.quantita,"prezzo":r.prezzo_snapshot,"subtot":round(r.prezzo_snapshot*r.quantita,2)} for r in c.righe],
            "totale": round(c.totale or 0, 2),
            "tipo_chiusura": c.tipo_chiusura or "",
            "status": c.status,
        })
    except Exception as e:
        return err(str(e))
    finally:
        db.close()


# ── Chiudi e paga (marca come pagato, non libera ancora) ──────────────────────
@app.post("/api/comande/<int:cid>/paga")
def paga_comanda(cid):
    db = get_db()
    try:
        c = db.get(Comanda, cid)
        if not c: return err("Non trovata", 404)
        d = body()
        c.status = "chiusa"
        c.tipo_chiusura = d.get("tipo_chiusura","scontrino")
        c.closed_at = datetime.utcnow()
        t = db.get(Tavolo, c.tavolo_id)
        if t:
            t.status = "chiesto_conto"   # visivamente "pagato, in attesa liberazione"
            t.pagato = True
        db.commit(); db.refresh(c)
        return ok({"comanda": c.to_dict(), "tavolo": t.to_dict() if t else None})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ── Libera tavolo (dopo pagamento) ────────────────────────────────────────────
@app.post("/api/tavoli/<int:tid>/libera")
def libera_tavolo(tid):
    db = get_db()
    try:
        t = db.get(Tavolo, tid)
        if not t: return err("Non trovato", 404)
        if not t.pagato:
            return err("Il conto non è ancora stato pagato")
        t.status = "disponibile"
        t.coperti_attuali = 0
        t.turno_occupato_id = None
        t.tavolo_unito_id = None
        t.pagato = False
        db.commit(); db.refresh(t)
        return ok(t.to_dict())
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()




# ── Ripristina tavoli — rimuove tutti i link orfani ───────────────────────────
@app.post("/api/sale/<int:sid>/ripristina-tavoli")
def ripristina_tavoli(sid):
    """Forza il ripristino di tutti i tavoli disponibili ma con link rimasti."""
    db = get_db()
    try:
        tavoli = db.query(Tavolo).filter(Tavolo.sala_id == sid).all()
        fixed = 0
        for t in tavoli:
            changed = False
            # If tavolo is disponibile but has a stale union link → clear it
            if t.tavolo_unito_id and t.status == "disponibile":
                t.tavolo_unito_id = None
                changed = True
            # If partner no longer exists → clear link
            elif t.tavolo_unito_id:
                partner = db.get(Tavolo, t.tavolo_unito_id)
                if not partner:
                    t.tavolo_unito_id = None
                    changed = True
            if changed:
                fixed += 1
        db.commit()
        return ok({"fixed": fixed, "message": f"Ripristinati {fixed} tavoli"})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ── Dividi tavoli uniti ───────────────────────────────────────────────────────
@app.post("/api/tavoli/<int:tid>/dividi")
def dividi_tavolo(tid):
    """Divide un tavolo dalla sua unione, ripristinando entrambi come indipendenti."""
    db = get_db()
    try:
        t = db.get(Tavolo, tid)
        if not t: return err("Tavolo non trovato", 404)
        partner_id = t.tavolo_unito_id
        # Reset this table
        t.tavolo_unito_id = None
        # Reset partner too
        if partner_id:
            partner = db.get(Tavolo, partner_id)
            if partner:
                partner.tavolo_unito_id = None
                # If partner was secondary (attesa_ordine or riservato without comanda),
                # restore to appropriate status
                if partner.status in ("attesa_ordine", "riservato"):
                    # Check if partner has open comanda or prenotazione
                    open_comanda = db.query(Comanda).filter(
                        Comanda.tavolo_id == partner.id,
                        Comanda.status == "aperta"
                    ).first()
                    pren = db.query(Prenotazione).filter(
                        Prenotazione.tavolo_id == partner.id,
                        Prenotazione.confermata == True
                    ).first()
                    if not open_comanda and not pren:
                        partner.status = "disponibile"
                        partner.coperti_attuali = 0
        db.commit()
        return ok({"detail": "Divisione completata", "tavolo": t.to_dict()})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════════════════════
# RANGHI
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/sale/<int:sid>/ranghi")
def get_ranghi(sid):
    db = get_db()
    try:
        return ok([r.to_dict() for r in db.query(Rango).filter(Rango.sala_id==sid).all()])
    finally: db.close()

@app.post("/api/sale/<int:sid>/ranghi")
def create_rango(sid):
    db = get_db()
    try:
        d = body()
        r = Rango(sala_id=sid, nome=d.get("nome","Rango 1"),
                  colore=d.get("colore","#3b82f6"), cameriere=d.get("cameriere"))
        db.add(r); db.commit(); db.refresh(r)
        return ok(r.to_dict(), 201)
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()

@app.delete("/api/ranghi/<int:rid>")
def delete_rango(rid):
    db = get_db()
    try:
        r = db.get(Rango, rid)
        if not r: return err("Non trovato", 404)
        db.delete(r); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()




# ── Check-in prenotazione ──────────────────────────────────────────────────────
@app.post("/api/prenotazioni/<int:pid>/checkin")
def checkin_prenotazione(pid):
    """Effettua il check-in: marca la prenotazione come arrivata e mette il tavolo in attesa_ordine."""
    db = get_db()
    try:
        p = db.get(Prenotazione, pid)
        if not p: return err("Prenotazione non trovata", 404)
        if p.tavolo_id:
            t = db.get(Tavolo, p.tavolo_id)
            if t:
                t.status = "attesa_ordine"
                t.coperti_attuali = p.coperti
        db.commit()
        return ok({"detail": "Check-in effettuato", "tavolo_status": "attesa_ordine"})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()

# ═══════════════════════════════════════════════════════════════════════════
# STATS
# ═══════════════════════════════════════════════════════════════════════════
@app.get("/api/sala/<int:sid>/stats")
def get_stats(sid):
    db = get_db()
    try:
        tavoli = db.query(Tavolo).filter(Tavolo.sala_id==sid).all()
        tav_ids = [t.id for t in tavoli]
        # Filter fatturato by date if provided
        data_filter = request.args.get("data")
        q_fat = db.query(Comanda).filter(
            Comanda.tavolo_id.in_(tav_ids),
            Comanda.status == "chiusa"
        )
        if data_filter:
            # Filter by closed_at date
            from sqlalchemy import func
            q_fat = q_fat.filter(func.date(Comanda.closed_at) == data_filter)
        fatturato = sum(c.totale or 0 for c in q_fat.all())
        return ok({
            "totale_tavoli": len(tavoli),
            "liberi":  sum(1 for t in tavoli if t.status=="disponibile"),
            "occupati":sum(1 for t in tavoli if t.status=="occupato"),
            "prenotati":sum(1 for t in tavoli if t.status=="riservato"),
            "chiesto_conto":sum(1 for t in tavoli if t.status=="chiesto_conto"),
            "fatturato_oggi": round(fatturato, 2),
        })
    finally: db.close()

@app.get("/api/dashboard")
def dashboard():
    db = get_db()
    try:
        n_outlets  = db.query(Outlet).filter(Outlet.attivo==True).count()
        n_voci     = db.query(VoceMenu).filter(VoceMenu.attivo==True).count()
        n_pren_oggi= db.query(Prenotazione).filter(
            Prenotazione.data==datetime.utcnow().strftime("%Y-%m-%d")).count()
        fatturato_oggi = db.query(Comanda).filter(
            Comanda.status=="chiusa",
            Comanda.closed_at >= datetime.utcnow().strftime("%Y-%m-%d")).all()
        tot = round(sum(c.totale for c in fatturato_oggi), 2)
        return ok({"outlets":n_outlets,"voci_menu":n_voci,
                   "prenotazioni_oggi":n_pren_oggi,"fatturato_oggi":tot})
    finally: db.close()


if __name__ == "__main__":
    app.run(port=8000, debug=True)


# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════
from datetime import timedelta
from sqlalchemy.orm import joinedload
from .models import Ruolo as RuoloModel, PermessoRuolo, Utente, Sessione, PAGINE_SISTEMA
from .auth import get_current_user, require_auth, require_admin


def _q_utenti(db):
    return db.query(Utente).options(
        joinedload(Utente.ruolo).joinedload(RuoloModel.permessi)
    )

def _q_ruoli(db):
    return db.query(RuoloModel).options(joinedload(RuoloModel.permessi))

def _build_permessi(ruolo):
    if not ruolo: return {}
    if ruolo.is_admin: return {p[0]: "completa" for p in PAGINE_SISTEMA}
    return {p.pagina: p.accesso for p in ruolo.permessi}


def _seed_auth():
    db = get_db()
    try:
        if db.query(RuoloModel).count() == 0:
            admin_role = RuoloModel(nome="Admin", descrizione="Accesso completo", is_admin=True)
            db.add(admin_role); db.flush()
            if db.query(Utente).count() == 0:
                u = Utente(username="admin", email="admin@outlet.local",
                           full_name="Amministratore", ruolo_id=admin_role.id, attivo=True)
                u.set_password("admin123")
                db.add(u)
            db.commit()
            print("✅ Utente admin creato — username: admin / password: admin123")
    except Exception as e:
        db.rollback(); print(f"⚠️  Seed auth: {e}")
    finally:
        db.close()

with app.app_context():
    _seed_auth()


@app.post("/api/auth/login")
def login():
    db = get_db()
    try:
        d = body()
        username = (d.get("username") or "").strip()
        password = d.get("password") or ""
        if not username or not password:
            return err("Username e password obbligatori")
        u = _q_utenti(db).filter(Utente.username == username).first()
        if not u or not u.attivo:
            return err("Credenziali non valide", 401)
        if not u.check_password(password):
            return err("Credenziali non valide", 401)
        token = Sessione.new_token()
        sess = Sessione(token=token, utente_id=u.id,
                        expires_at=datetime.utcnow() + timedelta(hours=24))
        db.add(sess)
        u.last_login = datetime.utcnow()
        db.commit()
        permessi = _build_permessi(u.ruolo)
        user_dict = u.to_dict()
        return ok({"token": token, "user": user_dict, "permessi": permessi})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.post("/api/auth/logout")
def logout():
    db = get_db()
    try:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:].strip()
            db.query(Sessione).filter(Sessione.token == token).delete()
            db.commit()
        return ok({"detail": "Logout effettuato"})
    finally:
        db.close()


@app.get("/api/auth/me")
@require_auth
def me():
    db = get_db()
    try:
        from flask import g
        u = _q_utenti(db).filter(Utente.id == g.user.id).first()
        if not u: return err("Non trovato", 404)
        permessi = _build_permessi(u.ruolo)
        user_dict = u.to_dict()
        return ok({"user": user_dict, "permessi": permessi})
    finally:
        db.close()


# ── Ruoli ─────────────────────────────────────────────────────────────────────
@app.get("/api/ruoli")
@require_auth
def get_ruoli():
    db = get_db()
    try:
        inc = request.args.get("include_permessi") == "1"
        ruoli = _q_ruoli(db).order_by(RuoloModel.nome).all()
        result = [r.to_dict(include_permessi=inc) for r in ruoli]
        return ok(result)
    finally:
        db.close()


@app.get("/api/ruoli/<int:rid>")
@require_auth
def get_ruolo(rid):
    db = get_db()
    try:
        r = _q_ruoli(db).filter(RuoloModel.id == rid).first()
        if not r: return err("Non trovato", 404)
        result = r.to_dict(include_permessi=True)
        return ok(result)
    finally:
        db.close()


@app.post("/api/ruoli")
@require_admin
def create_ruolo():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("nome obbligatorio")
        r = RuoloModel(nome=d["nome"], descrizione=d.get("descrizione",""), is_admin=d.get("is_admin",False))
        db.add(r); db.flush()
        for pagina, *_ in PAGINE_SISTEMA:
            acc = d.get("permessi", {}).get(pagina, "nascosta")
            db.add(PermessoRuolo(ruolo_id=r.id, pagina=pagina, accesso=acc))
        db.commit()
        r2 = _q_ruoli(db).filter(RuoloModel.id == r.id).first()
        result = r2.to_dict(include_permessi=True)
        return ok(result, 201)
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.put("/api/ruoli/<int:rid>")
@require_admin
def update_ruolo(rid):
    db = get_db()
    try:
        r = db.get(RuoloModel, rid)
        if not r: return err("Non trovato", 404)
        d = body()
        for k in ("nome","descrizione","is_admin"):
            if k in d: setattr(r, k, d[k])
        if "permessi" in d:
            db.query(PermessoRuolo).filter(PermessoRuolo.ruolo_id == rid).delete()
            db.flush()
            for pagina, *_ in PAGINE_SISTEMA:
                acc = d["permessi"].get(pagina, "nascosta")
                db.add(PermessoRuolo(ruolo_id=rid, pagina=pagina, accesso=acc))
        db.commit()
        r2 = _q_ruoli(db).filter(RuoloModel.id == rid).first()
        result = r2.to_dict(include_permessi=True)
        return ok(result)
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.delete("/api/ruoli/<int:rid>")
@require_admin
def delete_ruolo(rid):
    db = get_db()
    try:
        r = db.get(RuoloModel, rid)
        if not r: return err("Non trovato", 404)
        if r.is_admin: return err("Il ruolo Admin non può essere eliminato")
        db.delete(r); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.get("/api/pagine-sistema")
@require_auth
def pagine_sistema():
    return ok([{"id": p[0], "label": p[1], "sezione": p[2]} for p in PAGINE_SISTEMA])


# ── Stampanti ─────────────────────────────────────────────────────────────────
@app.get("/api/stampanti")
def get_stampanti():
    db = get_db()
    try:
        q = db.query(Stampante)
        if request.args.get("outlet_id"):
            q = q.filter(Stampante.outlet_id==int(request.args["outlet_id"]))
        if request.args.get("tipo"):
            q = q.filter(Stampante.tipo==request.args["tipo"])
        return ok([s.to_dict() for s in q.all()])
    finally: db.close()


@app.post("/api/stampanti")
def create_stampante():
    db = get_db()
    try:
        d = body()
        s = Stampante(nome=d["nome"], ip_address=d.get("ip_address",""),
                      protocollo=d.get("protocollo","epson"),
                      tipo=d.get("tipo","produzione"),
                      outlet_id=d.get("outlet_id") or None,
                      attiva=d.get("attiva",True))
        db.add(s); db.commit(); db.refresh(s)
        return ok(s.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.put("/api/stampanti/<int:sid>")
def update_stampante(sid):
    db = get_db()
    try:
        s = db.get(Stampante, sid)
        if not s: return err("Non trovata", 404)
        d = body()
        for k in ("nome","ip_address","protocollo","tipo","attiva"):
            if k in d: setattr(s, k, d[k])
        if "outlet_id" in d: s.outlet_id = d["outlet_id"] or None
        db.commit(); db.refresh(s)
        return ok(s.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/stampanti/<int:sid>")
def delete_stampante(sid):
    db = get_db()
    try:
        s = db.get(Stampante, sid)
        if not s: return err("Non trovata", 404)
        db.delete(s); db.commit()
        return ok({"deleted": sid})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.post("/api/voci-menu/<int:vid>/stampanti")
def set_voce_stampanti(vid):
    """Replace all printer links for a voce menu. Body: [{stampante_id, outlet_id, contesto}]"""
    db = get_db()
    try:
        items = body()  # list of {stampante_id, outlet_id, contesto}
        # Remove existing
        db.query(VoceMenuStampante).filter(VoceMenuStampante.voce_id==vid).delete()
        for item in items:
            db.add(VoceMenuStampante(voce_id=vid,
                                     stampante_id=int(item["stampante_id"]),
                                     outlet_id=int(item["outlet_id"]),
                                     contesto=item.get("contesto","reparto_produzione")))
        db.commit()
        return ok({"updated": len(items)})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.get("/api/voci-menu/<int:vid>/stampanti")
def get_voce_stampanti(vid):
    db = get_db()
    try:
        links = db.query(VoceMenuStampante).filter(VoceMenuStampante.voce_id==vid).all()
        return ok([{"stampante_id":l.stampante_id,"outlet_id":l.outlet_id,"contesto":l.contesto} for l in links])
    finally: db.close()


@app.get("/api/outlets/<int:oid>/next-comanda-numero")
def next_comanda_numero(oid):
    db = get_db()
    try:
        from sqlalchemy import func as sqlfunc
        outlet = db.get(Outlet, oid)
        # Build 2-letter prefix from outlet name (skip numbers/spaces)
        if outlet:
            letters = [c for c in outlet.nome if c.isalpha()]
            prefix = "".join(letters[:2]).upper() if len(letters) >= 2 else f"O{oid}"
        else:
            prefix = f"O{oid}"
        # Count comande for this outlet only
        count = db.query(sqlfunc.count(Comanda.id)).filter(Comanda.outlet_id==oid).scalar() or 0
        numero = f"{prefix}{count+1:03d}"
        return ok({"numero": numero, "prefix": prefix})
    finally: db.close()



@app.post("/api/stampa-reparti")
def stampa_reparti():
    """Return print groups: for each printer, list of items to print.
    Body: {outlet_id, righe:[{voce_id,nome,qty,prezzo}], contesto:'reparto_produzione'|'chiusura_comanda'}
    """
    db = get_db()
    try:
        d = body()
        outlet_id = d.get("outlet_id")
        righe     = d.get("righe", [])
        contesto  = d.get("contesto", "reparto_produzione")

        # For each voce, find assigned printers for this outlet+contesto
        printer_groups = {}  # stampante_id -> {stampante_info, voci:[]}

        for riga in righe:
            voce_id = riga.get("voce_id")
            if not voce_id:
                continue
            # Find printer links for this voce + outlet + contesto
            links = db.query(VoceMenuStampante).filter(
                VoceMenuStampante.voce_id   == int(voce_id),
                VoceMenuStampante.outlet_id == int(outlet_id),
                VoceMenuStampante.contesto  == contesto,
            ).all()

            for link in links:
                sid = link.stampante_id
                if sid not in printer_groups:
                    s = db.get(Stampante, sid)
                    if s and s.attiva:
                        printer_groups[sid] = {
                            "stampante": s.to_dict(),
                            "voci": []
                        }
                if sid in printer_groups:
                    printer_groups[sid]["voci"].append({
                        "voce_id": voce_id,
                        "nome":    riga.get("nome",""),
                        "qty":     riga.get("qty", 1),
                        "prezzo":  riga.get("prezzo", 0),
                        "note":    riga.get("note",""),
                    })

        return ok(list(printer_groups.values()))
    except Exception as e:
        return err(str(e))
    finally:
        db.close()


# ── Utenti ────────────────────────────────────────────────────────────────────


@app.get("/api/voci-menu/<int:vid>/monitor")
def get_voce_monitor(vid):
    """Get monitor links for a voce menu."""
    db = get_db()
    try:
        links = db.query(VoceMenuMonitor).filter(VoceMenuMonitor.voce_id==vid).all()
        return ok([{"monitor_id":l.monitor_id,"tutti_monitor":bool(l.tutti_monitor)} for l in links])
    finally: db.close()


@app.post("/api/voci-menu/<int:vid>/monitor")
def set_voce_monitor(vid):
    """Replace monitor links for a voce menu.
    Body: [{monitor_id: int|null, tutti_monitor: bool}]
    """
    db = get_db()
    try:
        items = body()
        db.query(VoceMenuMonitor).filter(VoceMenuMonitor.voce_id==vid).delete()
        for item in items:
            db.add(VoceMenuMonitor(
                voce_id=vid,
                monitor_id=int(item["monitor_id"]) if item.get("monitor_id") else None,
                tutti_monitor=bool(item.get("tutti_monitor", False))
            ))
        db.commit()
        return ok({"updated": len(items)})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# ── Web Menu ───────────────────────────────────────────────────────────────────

@app.get("/api/web-menu")
def get_web_menus():
    db = get_db()
    try:
        q = db.query(WebMenu)
        if request.args.get("outlet_id"):
            q = q.filter(WebMenu.outlet_id==int(request.args["outlet_id"]))
        return ok([m.to_dict() for m in q.all()])
    finally: db.close()


@app.post("/api/web-menu")
def create_web_menu():
    db = get_db()
    try:
        d = body()
        nome = (d.get("nome") or "").strip()
        if not nome: return err("Nome obbligatorio")
        import re, uuid
        base = re.sub(r"[^a-z0-9]+", "-", nome.lower()).strip("-")
        slug = base + "-" + uuid.uuid4().hex[:6]
        m = WebMenu(nome=nome, slug=slug,
                    outlet_id=int(d["outlet_id"]) if d.get("outlet_id") else None,
                    titolo=d.get("titolo",""), sottotitolo=d.get("sottotitolo",""),
                    logo_url=d.get("logo_url",""),
                    colore_primario=d.get("colore_primario","#204769"),
                    colore_sfondo=d.get("colore_sfondo","#f8f9fa"),
                    colore_testo=d.get("colore_testo","#1a1a2a"),
                    colore_card=d.get("colore_card","#ffffff"),
                    font_famiglia=d.get("font_famiglia","'Inter','Segoe UI',sans-serif"),
                    mostra_prezzi=bool(d.get("mostra_prezzi",True)),
                    mostra_allergeni=bool(d.get("mostra_allergeni",True)),
                    nota_piede=d.get("nota_piede",""),
                    attivo=bool(d.get("attivo",True)),
                    data_dal=d.get("data_dal"),
                    data_al=d.get("data_al"),
                    servizio=d.get("servizio","Tutti"))
        db.add(m); db.commit(); db.refresh(m)
        return ok(m.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.put("/api/web-menu/<int:mid>")
def update_web_menu(mid):
    db = get_db()
    try:
        m = db.get(WebMenu, mid)
        if not m: return err("Non trovato", 404)
        d = body()
        for k in ("nome","titolo","sottotitolo","logo_url","colore_primario","colore_sfondo",
                   "colore_testo","colore_card","font_famiglia","mostra_prezzi","mostra_allergeni",
                   "nota_piede","attivo","data_dal","data_al","servizio"):
            if k in d: setattr(m, k, d[k])
        if "outlet_id" in d:
            m.outlet_id = int(d["outlet_id"]) if d["outlet_id"] else None
        if "voci" in d:
            for v in m.voci: db.delete(v)
            db.flush()
            for i, vd in enumerate(d["voci"]):
                db.add(WebMenuVoce(web_menu_id=mid,
                    voce_menu_id=int(vd["voce_menu_id"]) if vd.get("voce_menu_id") else None,
                    categoria=vd.get("categoria",""), nome=vd.get("nome",""),
                    descrizione=vd.get("descrizione",""), prezzo=vd.get("prezzo"),
                    allergeni=vd.get("allergeni",""), etichette=vd.get("etichette",""),
                    immagine_url=vd.get("immagine_url",""), ordine=i))
        db.commit(); db.refresh(m)
        return ok(m.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/web-menu/<int:mid>")
def delete_web_menu(mid):
    db = get_db()
    try:
        m = db.get(WebMenu, mid)
        if not m: return err("Non trovato", 404)
        db.delete(m); db.commit()
        return ok({"deleted": mid})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.get("/menu/<slug>")
def public_web_menu(slug):
    """Public HTML web menu page."""
    db = get_db()
    try:
        m = db.query(WebMenu).filter(WebMenu.slug==slug, WebMenu.attivo==True).first()
        if not m:
            return ("<html><body style='font-family:sans-serif;text-align:center;padding:80px'>"
                    "<h2>Menu non trovato</h2></body></html>"), 404, {"Content-Type":"text/html"}
        data = m.to_dict()
        cp = data["colore_primario"]; cs = data["colore_sfondo"]
        ct = data["colore_testo"];   cc = data["colore_card"]
        ff = data["font_famiglia"]
        cats = {}
        for v in data["voci"]:
            cats.setdefault(v["categoria"], []).append(v)
        cat_html = ""
        for cat, voci in cats.items():
            items = ""
            for v in voci:
                price = ""
                if data["mostra_prezzi"] and v.get("prezzo") is not None:
                    price = '<span class="price">€%.2f</span>' % v["prezzo"]
                desc = ('<p class="desc">%s</p>' % v["descrizione"]) if v.get("descrizione") else ""
                img  = ('<img src="%s" alt="%s" class="item-img"/>' % (v["immagine_url"], v["nome"])) if v.get("immagine_url") else ""
                badges = ""
                for tag in (v.get("etichette") or "").split(","):
                    tag = tag.strip()
                    if tag: badges += '<span class="badge">%s</span>' % tag
                allg = ""
                if data["mostra_allergeni"] and v.get("allergeni"):
                    allg = '<div class="allg">🌾 %s</div>' % v["allergeni"]
                items += ('<div class="item">%s<div class="item-body">'
                          '<div class="item-top"><span class="item-name">%s</span>%s</div>'
                          '%s<div class="item-foot">%s%s</div></div></div>' % (img, v["nome"], price, desc, badges, allg))
            cat_html += '<div class="cat-section"><h2 class="cat-title">%s</h2><div class="items">%s</div></div>' % (cat, items)
        logo_html = ('<img src="%s" alt="logo" class="logo"/>' % data["logo_url"]) if data.get("logo_url") else ""
        piede = ('<div class="footer-note">%s</div>' % data["nota_piede"]) if data.get("nota_piede") else ""
        sub   = ("<p>%s</p>" % data["sottotitolo"]) if data.get("sottotitolo") else ""
        title = data.get("titolo") or data["nome"]
        css = ("*{margin:0;padding:0;box-sizing:border-box}"
               "body{font-family:%s;background:%s;color:%s;min-height:100vh}"
               ".hero{background:%s;color:#fff;padding:32px 20px 24px;text-align:center}"
               ".logo{max-height:80px;max-width:200px;object-fit:contain;margin-bottom:14px;border-radius:8px}"
               ".hero h1{font-size:clamp(22px,5vw,36px);font-weight:900;margin-bottom:6px}"
               ".hero p{font-size:14px;opacity:.85;max-width:500px;margin:0 auto}"
               ".menu-body{max-width:720px;margin:0 auto;padding:24px 16px 40px}"
               ".cat-section{margin-bottom:32px}"
               ".cat-title{font-size:20px;font-weight:800;color:%s;margin-bottom:14px;padding-bottom:6px;border-bottom:2.5px solid rgba(0,0,0,.12);text-transform:uppercase;letter-spacing:.5px}"
               ".items{display:flex;flex-direction:column;gap:12px}"
               ".item{background:%s;border-radius:14px;overflow:hidden;display:flex;box-shadow:0 1px 6px rgba(0,0,0,.07);border:1px solid rgba(0,0,0,.06)}"
               ".item-img{width:110px;height:110px;object-fit:cover;flex-shrink:0}"
               ".item-body{flex:1;padding:12px 14px;display:flex;flex-direction:column;gap:4px}"
               ".item-top{display:flex;align-items:flex-start;justify-content:space-between;gap:10px}"
               ".item-name{font-size:15px;font-weight:700;flex:1;line-height:1.3}"
               ".price{font-size:16px;font-weight:900;color:%s;white-space:nowrap;flex-shrink:0}"
               ".desc{font-size:12px;opacity:.7;line-height:1.5}"
               ".item-foot{display:flex;flex-wrap:wrap;gap:5px;align-items:center;margin-top:4px}"
               ".badge{background:%s22;color:%s;border:1px solid %s33;border-radius:20px;font-size:10px;font-weight:700;padding:2px 8px}"
               ".allg{font-size:10px;opacity:.5;font-style:italic}"
               ".footer-note{text-align:center;font-size:11px;opacity:.6;padding:20px 16px 40px;line-height:1.7;border-top:1px solid rgba(0,0,0,.08);margin-top:16px}"
               "@media(max-width:480px){.item-img{width:90px;height:90px}.item-name{font-size:14px}}"
               ) % (ff, cs, ct, cp, cp, cc, cp, cp, cp, cp)
        html = ("<!DOCTYPE html><html lang='it'><head>"
                "<meta charset='UTF-8'><meta name='viewport' content='width=device-width,initial-scale=1'>"
                "<title>%s</title><style>%s</style></head><body>"
                "<div class='hero'>%s<h1>%s</h1>%s</div>"
                "<div class='menu-body'>%s</div>%s</body></html>") % (title, css, logo_html, title, sub, cat_html, piede)
        return html, 200, {"Content-Type":"text/html; charset=utf-8"}
    finally: db.close()


# ── Monitor KDS ───────────────────────────────────────────────────────────────

@app.get("/api/monitor")
def get_monitor_list():
    db = get_db()
    try:
        q = db.query(Monitor)
        if request.args.get("outlet_id"):
            q = q.filter(Monitor.outlet_id==int(request.args["outlet_id"]))
        return ok([m.to_dict() for m in q.all()])
    finally: db.close()


@app.post("/api/monitor")
def create_monitor():
    db = get_db()
    try:
        d = body()
        nome = d.get("nome","").strip()
        if not nome: return err("Nome obbligatorio")
        # Auto-generate slug from nome + reparto
        import re, uuid
        base = re.sub(r"[^a-z0-9]+", "-", nome.lower() + "-" + d.get("reparto","cucina")).strip("-")
        slug = base + "-" + uuid.uuid4().hex[:6]
        m = Monitor(
            nome=nome,
            reparto=d.get("reparto","cucina"),
            outlet_id=int(d["outlet_id"]) if d.get("outlet_id") else None,
            slug=slug,
            colore_sfondo=d.get("colore_sfondo","#1a1a2e"),
            colore_testo=d.get("colore_testo","#ffffff"),
            colore_griglia=d.get("colore_griglia","#2a2a3e"),
            colore_header=d.get("colore_header","#ffffff"),
            attivo=d.get("attivo",True)
        )
        db.add(m); db.commit(); db.refresh(m)
        return ok(m.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.put("/api/monitor/<int:mid>")
def update_monitor(mid):
    db = get_db()
    try:
        m = db.get(Monitor, mid)
        if not m: return err("Non trovato", 404)
        d = body()
        for k in ("nome","reparto","colore_sfondo","colore_testo","colore_griglia","colore_header","attivo"):
            if k in d: setattr(m, k, d[k])
        if "outlet_id" in d:
            m.outlet_id = int(d["outlet_id"]) if d["outlet_id"] else None
        db.commit(); db.refresh(m)
        return ok(m.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/monitor/<int:mid>")
def delete_monitor(mid):
    db = get_db()
    try:
        m = db.get(Monitor, mid)
        if not m: return err("Non trovato", 404)
        db.delete(m); db.commit()
        return ok({"deleted": mid})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.get("/api/monitor/<slug>/live")
def monitor_live(slug):
    """Public endpoint (no auth) — returns active comande for a monitor slug.
    Used by the KDS display page.
    """
    db = get_db()
    try:
        m = db.query(Monitor).filter(Monitor.slug==slug).first()
        if not m: return err("Monitor non trovato", 404)
        # Load open comande for this outlet
        q = db.query(Comanda).filter(Comanda.status=="aperta")
        if m.outlet_id:
            q = q.filter(Comanda.outlet_id==m.outlet_id)
        # Only show comande that have been sent to monitor (inviato_monitor=True)
        comande = q.filter(Comanda.inviato_monitor == True).order_by(Comanda.created_at).all()
        from datetime import datetime as _dt
        result = []
        for c in comande:
            righe = []
            for r in c.righe:
                voce = db.get(VoceMenu, r.voce_id) if r.voce_id else None
                if voce and hasattr(voce, "monitor_links") and voce.monitor_links:
                    # Check if this voce is linked to this monitor or tutti_monitor
                    show = any(
                        lnk.tutti_monitor or lnk.monitor_id == m.id
                        for lnk in voce.monitor_links
                    )
                    if not show:
                        continue
                # No monitor_links → show on all monitors (default behaviour)
                righe.append({"nome":r.nome_snapshot,"qty":r.quantita,
                              "prezzo":r.prezzo_snapshot,"turno_idx":r.turno_idx or 0})
            if not righe: continue
            delta = _dt.utcnow() - c.created_at if c.created_at else None
            wait_min = int(delta.total_seconds()/60) if delta else 0
            result.append({
                "id": c.id, "numero": c.numero,
                "tavolo_id": c.tavolo_id, "tavolo_num": None,
                "coperti": c.coperti, "note": c.note,
                "created_at": str(c.created_at),
                "wait_min": wait_min, "righe": righe,
                "turno_corrente": c.turno_corrente or 0,
                "num_turni": max((r.turno_idx or 0 for r in c.righe), default=0) + 1,
            })
        # Fill tavolo numbers
        from sqlalchemy.orm import joinedload
        tavolo_map = {}
        for row in result:
            if row["tavolo_id"] and row["tavolo_id"] not in tavolo_map:
                t = db.get(Tavolo, row["tavolo_id"])
                tavolo_map[row["tavolo_id"]] = t.numero if t else "?"
            row["tavolo_num"] = tavolo_map.get(row["tavolo_id"], "—")
        return ok({"monitor": m.to_dict(), "comande": result,
                   "ora": _dt.utcnow().strftime("%H:%M:%S")})
    finally: db.close()





# -- Mobile Wallet (Apple + Google) ----------------------------------------

@app.get('/api/mobile-wallet-config')
def get_mobile_wallet_config():
    db = get_db()
    try:
        cfg = db.get(ConfigMobileWallet, 1)
        if not cfg:
            cfg = ConfigMobileWallet(id=1); db.add(cfg); db.commit(); db.refresh(cfg)
        return ok(cfg.to_dict())
    finally: db.close()


@app.put('/api/mobile-wallet-config')
def update_mobile_wallet_config():
    db = get_db()
    try:
        d = body()
        cfg = db.get(ConfigMobileWallet, 1)
        if not cfg:
            cfg = ConfigMobileWallet(id=1); db.add(cfg)
        text_fields = [
            'apple_team_id','apple_pass_type_id','apple_org_name',
            'apple_key_password','google_issuer_id','google_class_id',
            'apple_enabled','google_enabled',
        ]
        for k in text_fields:
            if k in d: setattr(cfg, k, d[k])
        # PEM/JSON fields: only update if not placeholder
        for k in ('apple_cert_pem','apple_key_pem','apple_wwdr_pem','google_service_account'):
            if d.get(k) and d[k] != '***': setattr(cfg, k, d[k])
        db.commit(); db.refresh(cfg)
        return ok(cfg.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.get('/api/mobile-wallet/apple/<int:wid>')
def generate_apple_pass(wid):
    """Generate .pkpass file for an Apple Wallet pass."""
    db = get_db()
    try:
        cfg = db.get(ConfigMobileWallet, 1)
        if not cfg or not cfg.apple_enabled:
            return err('Apple Wallet non configurato o non abilitato')
        w = db.get(Wallet, wid)
        if not w: return err('Wallet non trovato')
        c = w.cliente
        import zipfile, hashlib, json, io, tempfile, os
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.serialization import pkcs7
        from cryptography import x509
        # Build pass.json
        saldo_str = "EUR %.2f" % (w.saldo or 0)
        nome = ((c.nome + " " + (c.cognome or "")).strip()) if c else "Cliente"
        pass_json = {
            "formatVersion": 1,
            "passTypeIdentifier": cfg.apple_pass_type_id,
            "serialNumber": "wallet-" + str(w.id),
            "teamIdentifier": cfg.apple_team_id,
            "organizationName": cfg.apple_org_name or "Outlet Manager",
            "description": "Wallet " + w.etichetta,
            "foregroundColor": "rgb(255,255,255)",
            "backgroundColor": "rgb(45,90,123)",
            "labelColor": "rgb(200,220,240)",
            "storeCard": {
                "primaryFields": [{"key":"balance","label":"Saldo disponibile","value":saldo_str,"currencyCode":"EUR"}],
                "secondaryFields": [{"key":"name","label":"Intestatario","value":nome}],
                "auxiliaryFields": [{"key":"wallet","label":"Wallet","value":w.etichetta}],
                "backFields": [{"key":"token","label":"Token","value":w.token}]
            },
            "barcodes": [{"message":w.token,"format":"PKBarcodeFormatQR","messageEncoding":"iso-8859-1","altText":"QR Wallet"}],
        }
        if w.data_scadenza:
            pass_json["expirationDate"] = w.data_scadenza + "T23:59:59Z"
        pass_json_bytes = json.dumps(pass_json, ensure_ascii=False).encode("utf-8")
        # Create manifest (SHA1 hashes)
        manifest = {"pass.json": hashlib.sha1(pass_json_bytes).hexdigest()}
        # Minimal 1x1 transparent PNG for icon (required by Apple)
        import base64
        tiny_png = base64.b64decode('iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg==')
        manifest["icon.png"] = hashlib.sha1(tiny_png).hexdigest()
        manifest_bytes = json.dumps(manifest).encode("utf-8")
        # Sign manifest with Apple certificate
        cert = x509.load_pem_x509_certificate(cfg.apple_cert_pem.encode())
        key  = serialization.load_pem_private_key(
            cfg.apple_key_pem.encode(),
            password=cfg.apple_key_password.encode() if cfg.apple_key_password else None
        )
        wwdr = x509.load_pem_x509_certificate(cfg.apple_wwdr_pem.encode())
        signature = (
            pkcs7.PKCS7SignatureBuilder()
            .set_data(manifest_bytes)
            .add_signer(cert, key, hashes.SHA256())
            .add_certificate(wwdr)
            .sign(serialization.Encoding.DER, [pkcs7.PKCS7Options.DetachedSignature])
        )
        # Bundle into .pkpass (zip)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("pass.json",    pass_json_bytes)
            zf.writestr("manifest.json",manifest_bytes)
            zf.writestr("signature",    signature)
            zf.writestr("icon.png",     tiny_png)
        buf.seek(0)
        fname = "wallet-%d.pkpass" % w.id
        from flask import Response
        return Response(buf.getvalue(),
            mimetype="application/vnd.apple.pkpass",
            headers={'Content-Disposition': 'attachment; filename=' + fname})
    except ImportError:
        return err('Installa cryptography: pip install cryptography')
    except Exception as e: return err(str(e))
    finally: db.close()


@app.get('/api/mobile-wallet/google/<int:wid>')
def generate_google_wallet_url(wid):
    """Generate Google Wallet save URL (JWT signed)."""
    db = get_db()
    try:
        cfg = db.get(ConfigMobileWallet, 1)
        if not cfg or not cfg.google_enabled:
            return err('Google Wallet non configurato o non abilitato')
        w = db.get(Wallet, wid)
        if not w: return err('Wallet non trovato')
        c = w.cliente
        import json, time
        try:
            import jwt
        except ImportError:
            return err('Installa PyJWT: pip install PyJWT')
        sa = json.loads(cfg.google_service_account)
        nome = ((c.nome + " " + (c.cognome or "")).strip()) if c else "Cliente"
        saldo_str = "EUR %.2f" % (w.saldo or 0)
        issuer_id = cfg.google_issuer_id
        class_id  = cfg.google_class_id or (issuer_id + ".wallet_class")
        object_id = issuer_id + ".wallet_obj_" + str(w.id)
        now = int(time.time())
        # Generic pass object
        generic_object = {
            "id": object_id,
            "classId": class_id,
            "genericType": "GENERIC_TYPE_UNSPECIFIED",
            "hexBackgroundColor": "#2d5a7b",
            "logo": {"sourceUri": {"uri": "https://api.qrserver.com/v1/create-qr-code/?size=100x100&data=" + w.token}},
            "cardTitle": {"defaultValue": {"language": "it", "value": "Wallet Digitale"}},
            "subheader": {"defaultValue": {"language": "it", "value": nome}},
            "header": {"defaultValue": {"language": "it", "value": w.etichetta}},
            "textModulesData": [{"id":"saldo","header":"Saldo","body":saldo_str}],
            "barcode": {"type":"QR_CODE","value":w.token,"alternateText":"QR Wallet"},
            "state": "ACTIVE",
        }
        if w.data_scadenza:
            generic_object["validTimeInterval"] = {"end": {"date": w.data_scadenza}}
        payload = {
            "iss": sa["client_email"],
            "aud": "google",
            "typ": "savetowallet",
            "iat": now,
            "payload": {"genericObjects": [generic_object]},
            "origins": [],
        }
        token = jwt.encode(payload, sa["private_key"], algorithm="RS256")
        save_url = "https://pay.google.com/gp/v/save/" + (token if isinstance(token,str) else token.decode())
        return ok({"url": save_url})
    except Exception as e: return err(str(e))
    finally: db.close()

# -- Email Config -----------------------------------------------------------

@app.get('/api/email-config')
def get_email_config():
    db = get_db()
    try:
        cfg = db.get(ConfigEmail, 1)
        if not cfg:
            cfg = ConfigEmail(id=1); db.add(cfg); db.commit(); db.refresh(cfg)
        return ok(cfg.to_dict())
    finally: db.close()


@app.put('/api/email-config')
def update_email_config():
    db = get_db()
    try:
        d = body()
        cfg = db.get(ConfigEmail, 1)
        if not cfg:
            cfg = ConfigEmail(id=1); db.add(cfg)
        for k in ('smtp_host','smtp_port','smtp_user','smtp_from_email',
                  'smtp_from_name','use_tls','use_ssl','attivo'):
            if k in d: setattr(cfg, k, d[k])
        if d.get('smtp_password') and d['smtp_password'] != '***':
            cfg.smtp_password = d['smtp_password']
        db.commit(); db.refresh(cfg)
        return ok(cfg.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


def _send_email(cfg, to_email, to_name, subject, html_body):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    msg = MIMEMultipart('alternative')
    frm = (cfg.smtp_from_name or 'Outlet Manager') + ' <' + (cfg.smtp_from_email or '') + '>'
    msg['From'] = frm
    msg['To'] = (to_name + ' <' + to_email + '>') if to_name else to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))
    port = int(cfg.smtp_port or 587)
    if cfg.use_ssl:
        s = smtplib.SMTP_SSL(cfg.smtp_host, port, timeout=15)
    else:
        s = smtplib.SMTP(cfg.smtp_host, port, timeout=15)
        if cfg.use_tls: s.starttls()
    if cfg.smtp_user and cfg.smtp_password:
        s.login(cfg.smtp_user, cfg.smtp_password)
    s.sendmail(cfg.smtp_from_email, to_email, msg.as_string())
    s.quit()


@app.post('/api/email-config/test')
def test_email():
    db = get_db()
    try:
        cfg = db.get(ConfigEmail, 1)
        if not cfg or not cfg.attivo: return err('Configurazione email non attiva')
        d = body()
        to = d.get('to_email') or cfg.smtp_from_email
        html = (
            "<html><body style='font-family:sans-serif;padding:20px'>"
            "<h2>Test Email Outlet Manager</h2>"
            "<p>La configurazione SMTP funziona correttamente.</p>"
            "</body></html>"
        )
        _send_email(cfg, to, '', 'Test Email - Outlet Manager', html)
        return ok({'sent': True, 'to': to})
    except Exception as e: return err(str(e))
    finally: db.close()


@app.post('/api/email-config/send-wallet')
def send_wallet_email():
    db = get_db()
    try:
        cfg = db.get(ConfigEmail, 1)
        if not cfg or not cfg.attivo:
            return err('Servizio email non configurato. Vai in Generali > Configurazione Email.')
        d = body()
        wid = d.get('wallet_id')
        if not wid: return err('wallet_id obbligatorio')
        w = db.get(Wallet, int(wid))
        if not w: return err('Wallet non trovato')
        c = w.cliente
        if not c or not c.email: return err('Il cliente non ha email configurato')
        qr_url = (
            'https://api.qrserver.com/v1/create-qr-code/'
            '?size=200x200&data=' + w.token + '&bgcolor=ffffff'
        )
        scad = ('Valido fino al: ' + w.data_scadenza + '<br>') if w.data_scadenza else ''
        nome = (c.nome + ' ' + (c.cognome or '')).strip()
        saldo_fmt = '%.2f' % (w.saldo or 0.0)
        html = (
            "<html><body style='font-family:sans-serif;max-width:520px;margin:0 auto;padding:20px'>"
            "<div style='background:#2d5a7b;padding:20px;text-align:center'>"
            "<h1 style='color:white;margin:0'>Wallet Digitale</h1></div>"
            "<div style='background:#f8fafc;padding:24px;border:1px solid #e2e8f0'>"
            "<p>Gentile <b>" + nome + "</b>,</p>"
            "<p>Il tuo wallet <b>" + w.etichetta + "</b> e attivo.</p>"
            "<div style='text-align:center;margin:16px 0'>"
            "<img src='" + qr_url + "' style='width:180px;height:180px'/>"
            "<div style='font-size:28px;font-weight:900;color:#2d5a7b'>EUR " + saldo_fmt + "</div>"
            "</div>"
            "<p style='font-size:12px;color:#64748b'>" + scad + "Mostra il QR al pagamento.</p>"
            "</div></body></html>"
        )
        subj = 'Wallet ' + w.etichetta + ' - Saldo EUR ' + saldo_fmt
        _send_email(cfg, c.email, nome, subj, html)
        return ok({'sent': True, 'to': c.email})
    except Exception as e: return err(str(e))
    finally: db.close()

# -- Clienti ----------------------------------------------------------------

@app.get("/api/clienti")
def get_clienti():
    db = get_db()
    try:
        q = db.query(Cliente)
        if request.args.get("q"):
            s = "%" + request.args["q"] + "%"
            q = q.filter((Cliente.nome.ilike(s))|(Cliente.cognome.ilike(s))|(Cliente.email.ilike(s)))
        return ok([c.to_dict() for c in q.order_by(Cliente.cognome, Cliente.nome).all()])
    finally: db.close()


@app.post("/api/clienti")
def create_cliente():
    db = get_db()
    try:
        d = body()
        if not d.get("nome"): return err("Nome obbligatorio")
        c = Cliente(nome=d["nome"], cognome=d.get("cognome",""),
                    email=d.get("email",""), telefono=d.get("telefono",""),
                    note=d.get("note",""),
                    categoria_cliente_id=int(d["categoria_cliente_id"]) if d.get("categoria_cliente_id") else None)
        db.add(c); db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.put("/api/clienti/<int:cid>")
def update_cliente(cid):
    db = get_db()
    try:
        c = db.get(Cliente, cid)
        if not c: return err("Non trovato", 404)
        d = body()
        for k in ("nome","cognome","email","telefono","note"):
            if k in d: setattr(c, k, d[k])
        if "categoria_cliente_id" in d:
            c.categoria_cliente_id = int(d["categoria_cliente_id"]) if d["categoria_cliente_id"] else None
        db.commit(); db.refresh(c)
        return ok(c.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/clienti/<int:cid>")
def delete_cliente(cid):
    db = get_db()
    try:
        c = db.get(Cliente, cid)
        if not c: return err("Non trovato", 404)
        db.delete(c); db.commit()
        return ok({"deleted": cid})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


# -- Wallets -----------------------------------------------------------------

@app.get("/api/wallets")
def get_wallets():
    db = get_db()
    try:
        q = db.query(Wallet)
        if request.args.get("cliente_id"): q = q.filter(Wallet.cliente_id==int(request.args["cliente_id"]))
        if request.args.get("outlet_id"):  q = q.filter(Wallet.outlet_id==int(request.args["outlet_id"]))
        return ok([w.to_dict() for w in q.order_by(Wallet.created_at.desc()).all()])
    finally: db.close()


@app.get("/api/wallets/scan/<token>")
def scan_wallet(token):
    """Public endpoint: identify wallet by QR token."""
    db = get_db()
    try:
        w = db.query(Wallet).filter(Wallet.token==token).first()
        if not w: return err("Wallet non trovato", 404)
        if not w.attivo: return err("Wallet disattivato", 403)
        from datetime import datetime as _dt
        if w.data_scadenza and str(w.data_scadenza) < _dt.utcnow().strftime("%Y-%m-%d"):
            return err("Wallet scaduto", 403)
        data = w.to_dict()
        data["transazioni_recenti"] = [t.to_dict() for t in
            db.query(WalletTransazione).filter(WalletTransazione.wallet_id==w.id)
            .order_by(WalletTransazione.created_at.desc()).limit(10).all()]
        return ok(data)
    finally: db.close()


@app.post("/api/wallets")
def create_wallet():
    db = get_db()
    try:
        d = body()
        if not d.get("cliente_id"): return err("cliente_id obbligatorio")
        import secrets as _sec
        token = _sec.token_urlsafe(32)
        w = Wallet(cliente_id=int(d["cliente_id"]),
                   outlet_id=int(d["outlet_id"]) if d.get("outlet_id") else None,
                   etichetta=d.get("etichetta","Wallet"),
                   saldo=float(d.get("saldo_iniziale",0)),
                   token=token, attivo=True,
                   data_scadenza=d.get("data_scadenza"))
        db.add(w)
        if float(d.get("saldo_iniziale",0)) > 0:
            db.flush()
            db.add(WalletTransazione(wallet_id=w.id, tipo="ricarica",
                importo=float(d["saldo_iniziale"]), note="Ricarica iniziale"))
        db.commit(); db.refresh(w)
        return ok(w.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.put("/api/wallets/<int:wid>")
def update_wallet(wid):
    db = get_db()
    try:
        w = db.get(Wallet, wid)
        if not w: return err("Non trovato", 404)
        d = body()
        for k in ("etichetta","attivo","data_scadenza"):
            if k in d: setattr(w, k, d[k])
        db.commit(); db.refresh(w)
        return ok(w.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.delete("/api/wallets/<int:wid>")
def delete_wallet(wid):
    db = get_db()
    try:
        w = db.get(Wallet, wid)
        if not w: return err("Non trovato", 404)
        db.delete(w); db.commit()
        return ok({"deleted": wid})
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.post("/api/wallets/<int:wid>/ricarica")
def ricarica_wallet(wid):
    db = get_db()
    try:
        w = db.get(Wallet, wid)
        if not w: return err("Non trovato", 404)
        d = body()
        importo = float(d.get("importo",0))
        if importo <= 0: return err("Importo deve essere positivo")
        w.saldo = (w.saldo or 0) + importo
        t = WalletTransazione(wallet_id=wid, tipo="ricarica", importo=importo,
                               note=d.get("note","Ricarica"))
        db.add(t); db.commit(); db.refresh(w)
        return ok(w.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.post("/api/wallets/<int:wid>/paga")
def paga_wallet(wid):
    """Scalare dal wallet. Checks: active, not expired, sufficient balance."""
    db = get_db()
    try:
        w = db.get(Wallet, wid)
        if not w: return err("Wallet non trovato", 404)
        if not w.attivo: return err("Wallet disattivato", 403)
        from datetime import datetime as _dt
        if w.data_scadenza and str(w.data_scadenza) < _dt.utcnow().strftime("%Y-%m-%d"):
            return err("Wallet scaduto", 403)
        d = body()
        importo = float(d.get("importo",0))
        if importo <= 0: return err("Importo deve essere positivo")
        if (w.saldo or 0) < importo: return err(f"Credito insufficiente (disponibile: €{w.saldo:.2f})", 402)
        w.saldo = (w.saldo or 0) - importo
        t = WalletTransazione(wallet_id=wid, tipo="pagamento", importo=-importo,
                               note=d.get("note","Pagamento comanda"),
                               comanda_id=int(d["comanda_id"]) if d.get("comanda_id") else None)
        db.add(t); db.commit(); db.refresh(w)
        return ok(w.to_dict())
    except Exception as e: db.rollback(); return err(str(e))
    finally: db.close()


@app.get("/api/wallets/<int:wid>/movimenti")
def get_movimenti(wid):
    db = get_db()
    try:
        rows = db.query(WalletTransazione).filter(WalletTransazione.wallet_id==wid)\
                  .order_by(WalletTransazione.created_at.desc()).all()
        return ok([r.to_dict() for r in rows])
    finally: db.close()


@app.get("/api/utenti")
@require_admin
def get_utenti():
    db = get_db()
    try:
        utenti = _q_utenti(db).order_by(Utente.full_name).all()
        result = [u.to_dict() for u in utenti]
        return ok(result)
    finally:
        db.close()


@app.post("/api/utenti")
@require_admin
def create_utente():
    db = get_db()
    try:
        d = body()
        if not d.get("username"): return err("username obbligatorio")
        if not d.get("password"): return err("password obbligatoria")
        if db.query(Utente).filter(Utente.username == d["username"]).first():
            return err("Username già in uso")
        u = Utente(username=d["username"], email=d.get("email",""),
                   full_name=d.get("full_name",""), ruolo_id=d.get("ruolo_id"), attivo=True)
        u.set_password(d["password"])
        db.add(u); db.commit()
        u2 = _q_utenti(db).filter(Utente.id == u.id).first()
        result = u2.to_dict()
        return ok(result, 201)
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.put("/api/utenti/<int:uid>")
@require_admin
def update_utente(uid):
    db = get_db()
    try:
        u = db.get(Utente, uid)
        if not u: return err("Non trovato", 404)
        d = body()
        for k in ("username","email","full_name","ruolo_id","attivo"):
            if k in d: setattr(u, k, d[k])
        if d.get("password"):
            u.set_password(d["password"])
        db.commit()
        u2 = _q_utenti(db).filter(Utente.id == uid).first()
        result = u2.to_dict()
        return ok(result)
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.delete("/api/utenti/<int:uid>")
@require_admin
def delete_utente(uid):
    db = get_db()
    try:
        u = _q_utenti(db).filter(Utente.id == uid).first()
        if not u: return err("Non trovato", 404)
        if u.ruolo and u.ruolo.is_admin:
            altri = db.query(Utente).join(RuoloModel).filter(
                RuoloModel.is_admin == True, Utente.id != uid, Utente.attivo == True).count()
            if altri == 0:
                return err("Impossibile eliminare l'unico amministratore attivo")
        db.delete(u); db.commit()
        return ok({"detail": "eliminato"})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.post("/api/utenti/<int:uid>/reset-password")
@require_admin
def reset_password(uid):
    db = get_db()
    try:
        u = db.get(Utente, uid)
        if not u: return err("Non trovato", 404)
        d = body()
        if not d.get("password"): return err("password obbligatoria")
        u.set_password(d["password"])
        db.commit()
        return ok({"detail": "Password aggiornata"})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# PRENOTAZIONI — Capacità turno + Export
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/turni/<int:tid>/capacita")
def get_capacita_turno(tid):
    """Coperti prenotati per turno+data (solo prenotazioni con turno_id==tid O senza turno)."""
    db = get_db()
    try:
        data = request.args.get("data")
        turno = db.get(Turno, tid)
        if not turno: return err("Turno non trovato", 404)

        sala_id = request.args.get("sala_id")

        q = db.query(Prenotazione).filter(
            Prenotazione.outlet_id == turno.outlet_id,
            Prenotazione.confermata == True,
        )
        if data:
            q = q.filter(Prenotazione.data == data)
        if sala_id:
            q = q.filter(Prenotazione.sala_id == int(sala_id))

        # Conta solo le prenotazioni che appartengono a questo turno:
        # quelle con turno_id == tid OPPURE quelle senza turno ma con orario
        # ricadente nella finestra del turno
        from datetime import time as dtime
        def in_finestra(orario_str):
            try:
                h, m = map(int, orario_str.split(":"))
                oh, om = map(int, turno.ora_inizio.split(":"))
                eh, em = map(int, turno.ora_fine.split(":"))
                t = h*60+m; ts = oh*60+om; te = eh*60+em
                return ts <= t <= te
            except:
                return False

        prenotazioni = q.all()
        prenotati = sum(
            p.coperti for p in prenotazioni
            if (p.turno_id is not None and p.turno_id == tid) or
               (p.turno_id is None and in_finestra(p.orario or ""))
        )
        max_cap = turno.copertura_max or 0
        disponibili = max(0, max_cap - prenotati) if max_cap > 0 else None

        return ok({
            "turno_id":     tid,
            "data":         data,
            "copertura_max":max_cap,
            "prenotati":    prenotati,
            "disponibili":  disponibili,
            "al_completo":  (disponibili == 0) if max_cap > 0 else False,
        })
    finally:
        db.close()


@app.get("/api/prenotazioni/heatmap")
def prenotazioni_heatmap():
    """Returns booking counts per day for a given year-month (or full year).
    Query: ?year=2026&month=4  (month optional → full year)
    Returns: {"YYYY-MM-DD": {"count": N, "coperti": N}, ...}
    """
    db = get_db()
    try:
        year  = int(request.args.get("year",  __import__("datetime").datetime.utcnow().year))
        month = request.args.get("month")
        q = db.query(Prenotazione)
        if month:
            from sqlalchemy import extract
            q = q.filter(extract("year",  Prenotazione.data.cast(__import__("sqlalchemy").Date)) == year,
                         extract("month", Prenotazione.data.cast(__import__("sqlalchemy").Date)) == int(month))
        else:
            from sqlalchemy import extract
            q = q.filter(extract("year",  Prenotazione.data.cast(__import__("sqlalchemy").Date)) == year)
        result = {}
        for p in q.all():
            d = str(p.data)
            if d not in result:
                result[d] = {"count": 0, "coperti": 0}
            result[d]["count"]   += 1
            result[d]["coperti"] += p.coperti or 0
        return ok(result)
    except Exception as e:
        return err(str(e))
    finally:
        db.close()


@app.get("/api/prenotazioni/export")
def export_prenotazioni():
    """Esporta prenotazioni come XLSX, JSON (per PDF) o CSV fallback."""
    db = get_db()
    try:
        outlet_id = request.args.get("outlet_id")
        data      = request.args.get("data")
        turno_id  = request.args.get("turno_id")
        fmt       = request.args.get("format", "json")   # xlsx | json

        q = db.query(Prenotazione)
        if outlet_id: q = q.filter(Prenotazione.outlet_id == int(outlet_id))
        if data:      q = q.filter(Prenotazione.data == data)
        # Nessun filtro turno_id → tutte le prenotazioni del giorno
        # Con turno_id → solo quel turno + quelle senza turno nell'orario del turno
        if turno_id and turno_id.strip():
            tid = int(turno_id)
            turno_obj = db.get(Turno, tid)
            def in_win(orario_str):
                try:
                    h,m = map(int,(orario_str or "").split(":"))
                    oh,om = map(int,turno_obj.ora_inizio.split(":"))
                    eh,em = map(int,turno_obj.ora_fine.split(":"))
                    return oh*60+om <= h*60+m <= eh*60+em
                except: return False
            pren_all = q.order_by(Prenotazione.orario, Prenotazione.nome).all()
            pren = [p for p in pren_all if
                    (p.turno_id is not None and p.turno_id == tid) or
                    (p.turno_id is None and in_win(p.orario or ""))]
        else:
            pren = q.order_by(Prenotazione.orario, Prenotazione.nome).all()

        # Arricchisci con numero tavolo
        def enrich(p):
            d = p.to_dict()
            d["is_vip"] = getattr(p,"is_vip",False) or False
            if p.tavolo_id:
                t = db.get(Tavolo, p.tavolo_id)
                d["tavolo_numero"] = t.numero if t else None
            else:
                d["tavolo_numero"] = None
            return d

        rows = [enrich(p) for p in pren]

        # ── XLSX ─────────────────────────────────────────────────────────────
        if fmt == "xlsx":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                import io as _io

                wb = openpyxl.Workbook()
                ws = wb.active

                # Outlet info
                outlet = db.get(Outlet, int(outlet_id)) if outlet_id else None
                ws.title = "Prenotazioni"

                # Header info
                ws.merge_cells("A1:I1")
                ws["A1"] = f"{outlet.nome if outlet else ''} — Lista Ospiti {data or ''}"
                ws["A1"].font = Font(bold=True, size=13, color="1C2E4A")
                ws["A1"].alignment = Alignment(horizontal="left")

                ws.merge_cells("A2:I2")
                tot_pax = sum(r["coperti"] for r in rows)
                vip_cnt = sum(1 for r in rows if r["is_vip"])
                ws["A2"] = f"Totale: {len(rows)} prenotazioni · {tot_pax} coperti{' · ' + str(vip_cnt) + ' VIP' if vip_cnt else ''}"
                ws["A2"].font = Font(size=10, color="64748B")

                # Column headers
                HDR = ["VIP","Orario","Nome","Tel.","Email","Pax","Tavolo","Note","Confermata"]
                WIDTHS = [5,8,22,14,24,5,8,20,10]
                HDR_FILL = PatternFill("solid", fgColor="1C2E4A")
                HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
                border = Border(
                    bottom=Side(style="thin",color="E2E8F0"),
                    right=Side(style="thin",color="E2E8F0"),
                )

                for col, (h, w) in enumerate(zip(HDR, WIDTHS), 1):
                    cell = ws.cell(row=4, column=col, value=h)
                    cell.fill = HDR_FILL; cell.font = HDR_FONT
                    cell.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[cell.column_letter].width = w

                # Data rows
                GOLD = PatternFill("solid", fgColor="FFFBEB")
                ALT  = PatternFill("solid", fgColor="F8FAFC")

                for i, r in enumerate(rows):
                    row_n = i + 5
                    fill  = GOLD if r["is_vip"] else (ALT if i%2 else None)
                    vals  = [
                        "⭐" if r["is_vip"] else "",
                        r["orario"], r["nome"],
                        r.get("telefono") or "",
                        r.get("email") or "",
                        r["coperti"],
                        r.get("tavolo_numero") or "—",
                        r.get("note") or "",
                        "Sì" if r.get("confermata") else "No",
                    ]
                    for col, v in enumerate(vals, 1):
                        cell = ws.cell(row=row_n, column=col, value=v)
                        cell.border = border
                        cell.alignment = Alignment(vertical="center", wrap_text=col==8)
                        if fill: cell.fill = fill

                ws.row_dimensions[1].height = 22
                ws.row_dimensions[4].height = 18
                ws.freeze_panes = "A5"

                buf = _io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                from flask import Response as Resp
                return Resp(
                    buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=prenotazioni_{data or 'export'}.xlsx"}
                )
            except ImportError:
                pass  # fallback to JSON if openpyxl not installed

        # ── JSON (per stampa PDF) ─────────────────────────────────────────────
        from collections import defaultdict
        per_orario = defaultdict(list)
        for r in rows:
            per_orario[r["orario"]].append(r)

        return ok({
            "data": data,
            "totale_pren": len(rows),
            "totale_coperti": sum(r["coperti"] for r in rows),
            "per_orario": dict(sorted(per_orario.items())),
            "lista": rows,
        })
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# ALLOCAZIONE AUTOMATICA TAVOLI
# ══════════════════════════════════════════════════════════════════════════════
@app.post("/api/prenotazioni/alloca-tavoli")
def alloca_tavoli():
    """
    Algoritmo intelligente di allocazione tavoli per una data/turno.
    
    Strategia:
    1. Ordina prenotazioni per coperti DESC (prima le più grandi)
    2. Per ogni prenotazione, cerca il tavolo singolo con capienza più vicina (minimo spreco)
    3. Se nessun tavolo singolo è sufficiente, cerca la coppia di tavoli adiacenti
       (per numero di tavolo consecutivo) con capienza combinata sufficiente
    4. Aggiorna tavolo_id e tavolo_unito_id nel DB
    5. Restituisce il piano di allocazione con statistiche
    """
    db = get_db()
    try:
        d = body()
        outlet_id = d.get("outlet_id")
        sala_id   = d.get("sala_id")
        data      = d.get("data")
        turno_id  = d.get("turno_id")

        if not outlet_id or not data:
            return err("outlet_id e data obbligatori")

        # Prenotazioni senza tavolo assegnato per quella data/sala/turno
        q = db.query(Prenotazione).filter(
            Prenotazione.outlet_id == outlet_id,
            Prenotazione.data == data,
            Prenotazione.tavolo_id == None,
            Prenotazione.confermata == True,
        )
        if sala_id:
            q = q.filter(Prenotazione.sala_id == int(sala_id))
        if turno_id:
            q = q.filter(
                (Prenotazione.turno_id == int(turno_id)) |
                (Prenotazione.turno_id == None)
            )
        pren_da_allocare = sorted(q.all(), key=lambda p: p.coperti, reverse=True)

        if not pren_da_allocare:
            return ok({"message": "Nessuna prenotazione da allocare", "allocazioni": [], "non_allocate": []})

        # Tavoli disponibili nella sala
        tavoli_q = db.query(Tavolo)
        if sala_id:
            tavoli_q = tavoli_q.filter(Tavolo.sala_id == int(sala_id))
        tutti_tavoli = sorted(tavoli_q.all(), key=lambda t: t.numero)

        # Tavoli già prenotati in quel giorno nella stessa sala (escludi quelli già occupati)
        tavoli_occupati = set()
        pren_esist_q = db.query(Prenotazione).filter(
            Prenotazione.outlet_id == outlet_id,
            Prenotazione.data == data,
            Prenotazione.tavolo_id != None,
        )
        if sala_id:
            pren_esist_q = pren_esist_q.filter(Prenotazione.sala_id == int(sala_id))
        if turno_id:
            pren_esist_q = pren_esist_q.filter(
                (Prenotazione.turno_id == int(turno_id)) |
                (Prenotazione.turno_id == None)
            )
        for p in pren_esist_q.all():
            if p.tavolo_id: tavoli_occupati.add(p.tavolo_id)

        # Pre-clean: clear any stale tavolo_unito_id before fresh allocation
        for t in tutti_tavoli:
            if t.tavolo_unito_id and t.id not in tavoli_occupati:
                t.tavolo_unito_id = None
        db.flush()

        tavoli_liberi = [t for t in tutti_tavoli if t.id not in tavoli_occupati and not t.bloccato]

        allocazioni   = []
        non_allocate  = []
        usati         = set()  # id tavoli già usati in questa allocazione

        for pren in pren_da_allocare:
            coperti = pren.coperti
            disponibili = [t for t in tavoli_liberi if t.id not in usati]

            # ── Strategia 1: tavolo singolo con capienza >= coperti, minimo spreco
            candidati_singoli = [t for t in disponibili if t.capienza >= coperti]
            candidati_singoli.sort(key=lambda t: t.capienza - coperti)  # minimo spreco

            if candidati_singoli:
                tavolo = candidati_singoli[0]
                usati.add(tavolo.id)
                pren.tavolo_id = tavolo.id
                pren.tavolo_unito_id = None
                db.flush()
                allocazioni.append({
                    "prenotazione_id": pren.id,
                    "nome": pren.nome,
                    "coperti": coperti,
                    "tavolo": tavolo.numero,
                    "capienza_tavolo": tavolo.capienza,
                    "tipo": "singolo",
                    "spreco": tavolo.capienza - coperti,
                })
                continue

            # ── Strategia 2: gruppo di N tavoli adiacenti (2, 3, 4...)
            def num_val(t):
                try: return int(''.join(filter(str.isdigit, t.numero)))
                except: return 0

            # Sort available by number for adjacency search
            disp_sorted = sorted([t for t in disponibili if t.id not in usati], key=num_val)
            best_group  = None
            best_waste  = 9999

            # Try groups of increasing size starting from 2
            for size in range(2, len(disp_sorted)+1):
                for start in range(len(disp_sorted)-size+1):
                    group = disp_sorted[start:start+size]
                    # Check consecutive adjacency (each pair within 1)
                    nums = [num_val(t) for t in group]
                    adiacenti = all(abs(nums[i+1]-nums[i])<=1 for i in range(len(nums)-1))
                    cap_tot = sum(t.capienza for t in group)
                    if cap_tot >= coperti:
                        waste = cap_tot - coperti
                        if waste < best_waste and (adiacenti or best_group is None):
                            best_group = group
                            best_waste = waste
                            break  # found optimal group of this size
                if best_group and best_waste == 0:
                    break  # perfect fit, stop

            if best_group:
                for tg in best_group:
                    usati.add(tg.id)
                # Primary = first table in group (lowest numero)
                t_prim = best_group[0]
                t_secondari = best_group[1:]
                pren.tavolo_id = t_prim.id
                # Store only first secondary for now (main FK); others via Tavolo.tavolo_unito_id
                pren.tavolo_unito_id = t_secondari[0].id if t_secondari else None
                db.flush()
                nomi = "+".join(t.numero for t in best_group)
                cap_tot2 = sum(t.capienza for t in best_group)
                allocazioni.append({
                    "prenotazione_id": pren.id,
                    "nome": pren.nome,
                    "coperti": coperti,
                    "tavolo": nomi,
                    "capienza_tavolo": cap_tot2,
                    "tipo": "unione" if len(best_group)>1 else "singolo",
                    "n_tavoli": len(best_group),
                    "spreco": cap_tot2 - coperti,
                    "_group_ids": [t.id for t in best_group],
                    "_prim_id": t_prim.id,
                })
                continue

            # ── Nessun tavolo disponibile
            non_allocate.append({
                "prenotazione_id": pren.id,
                "nome": pren.nome,
                "coperti": coperti,
                "motivo": f"Nessun tavolo libero con capienza >= {coperti}"
            })

        # Mark all allocated tavoli as "riservato" + set mutual tavolo_unito_id
        for alloc_item in allocazioni:
            group_ids = alloc_item.get("_group_ids", [])
            prim_id   = alloc_item.get("_prim_id")
            if not group_ids:
                # Singolo
                pid = alloc_item["prenotazione_id"]
                p2  = next((p for p in pren_da_allocare if p.id == pid), None)
                if p2 and p2.tavolo_id:
                    t2 = db.get(Tavolo, p2.tavolo_id)
                    if t2: t2.status = "riservato"
                continue
            # Group: primary stays active, secondaries point back to primary
            t_prim = db.get(Tavolo, prim_id) if prim_id else None
            if t_prim:
                t_prim.status = "riservato"
                t_prim.tavolo_unito_id = None   # primary has no back-link
            for tid in group_ids:
                if tid == prim_id: continue
                t_sec = db.get(Tavolo, tid)
                if t_sec:
                    t_sec.status = "riservato"
                    t_sec.tavolo_unito_id = prim_id  # secondary points to primary

        db.commit()

        return ok({
            "message": f"Allocati {len(allocazioni)} su {len(pren_da_allocare)} prenotazioni",
            "allocazioni": allocazioni,
            "non_allocate": non_allocate,
            "stats": {
                "totale": len(pren_da_allocare),
                "allocate": len(allocazioni),
                "non_allocate": len(non_allocate),
                "singoli": sum(1 for a in allocazioni if a["tipo"]=="singolo"),
                "unioni":  sum(1 for a in allocazioni if a["tipo"]=="unione"),
            }
        })
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


@app.delete("/api/prenotazioni/dealloca-tavoli")
def dealloca_tavoli():
    """Rimuove tutte le assegnazioni tavolo per una data/turno (reset allocazione)."""
    db = get_db()
    try:
        d = body()
        outlet_id = d.get("outlet_id")
        data      = d.get("data")
        turno_id  = d.get("turno_id")
        if not outlet_id or not data: return err("outlet_id e data obbligatori")

        q = db.query(Prenotazione).filter(
            Prenotazione.outlet_id == outlet_id,
            Prenotazione.data == data,
        )
        if turno_id:
            q = q.filter(
                (Prenotazione.turno_id == int(turno_id)) |
                (Prenotazione.turno_id == None)
            )
        n = 0
        tavoli_da_liberare = set()
        for p in q.all():
            if p.tavolo_id is not None:
                tavoli_da_liberare.add(p.tavolo_id)
                if p.tavolo_unito_id:
                    tavoli_da_liberare.add(p.tavolo_unito_id)
                p.tavolo_id = None
                p.tavolo_unito_id = None
                n += 1
        # Nuclear reset: clear ALL tavoli that were linked to these prenotazioni
        # and also scan the entire sala for orphan links
        all_tav_ids = set()
        for tid in tavoli_da_liberare:
            all_tav_ids.add(tid)

        # Find all tavoli in the affected sale
        sala_ids = set()
        if sala_id:
            sala_ids.add(int(sala_id))
        else:
            # Get all sale for this outlet
            from sqlalchemy import text as sqlt2
            rows = db.execute(sqlt2(
                "SELECT id FROM sale WHERE outlet_id=:oid"
            ), {"oid": int(outlet_id)}).fetchall()
            sala_ids = {r[0] for r in rows}

        # Reset ALL tavoli in affected sale that have tavolo_unito_id set
        for sid_clean in sala_ids:
            linked_tavoli = db.query(Tavolo).filter(
                Tavolo.sala_id == sid_clean,
                Tavolo.tavolo_unito_id != None
            ).all()
            for t in linked_tavoli:
                t.tavolo_unito_id = None
                if t.status == "riservato":
                    # Check if still has active pren (not being dealloca'd)
                    still_pren = db.query(Prenotazione).filter(
                        Prenotazione.tavolo_id == t.id,
                        Prenotazione.confermata == True,
                        Prenotazione.data == data
                    ).first()
                    if not still_pren:
                        t.status = "disponibile"
                        t.coperti_attuali = 0
        db.commit()
        return ok({"message": f"Rimossi {n} assegnazioni tavolo, liberati {len(tavoli_da_liberare)} tavoli", "rimossi": n})
    except Exception as e:
        db.rollback(); return err(str(e))
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD AVANZATA — dati real-time per grafici e monitoraggio
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/api/dashboard/live")
def dashboard_live():
    """
    Dati in tempo reale per la dashboard:
    - Stato tavoli per sala con timer occupazione
    - Andamento prenotazioni orario per orario
    - Ricavi turno corrente
    - Tavoli per stato
    """
    db = get_db()
    try:
        today = datetime.utcnow().strftime("%Y-%m-%d")
        now   = datetime.utcnow()

        # ── Sale e tavoli real-time ───────────────────────────────────────────
        sale_data = []
        for sala in db.query(Sala).filter(Sala.attiva == True).all():
            tavoli = db.query(Tavolo).filter(Tavolo.sala_id == sala.id).all()

            tavoli_detail = []
            for t in tavoli:
                # Comanda aperta per questo tavolo
                comanda = db.query(Comanda).filter(
                    Comanda.tavolo_id == t.id,
                    Comanda.status.in_(["aperta","chiesta_conto"])
                ).order_by(Comanda.created_at.desc()).first()

                # Prenotazione attiva oggi
                pren = db.query(Prenotazione).filter(
                    Prenotazione.tavolo_id == t.id,
                    Prenotazione.data == today,
                    Prenotazione.confermata == True,
                ).first()

                # Calcola durata occupazione in minuti
                durata_min = None
                if comanda and comanda.created_at:
                    delta = now - comanda.created_at
                    durata_min = int(delta.total_seconds() / 60)

                tavoli_detail.append({
                    "id":        t.id,
                    "numero":    t.numero,
                    "capienza":  t.capienza,
                    "status":    t.status,
                    "coperti":   t.coperti_attuali,
                    "hat_color": t.hat_color,
                    "durata_min":durata_min,
                    "totale_comanda": comanda.totale if comanda else 0,
                    "ha_prenotazione": pren is not None,
                    "nome_pren": pren.nome if pren else None,
                })

            # Statistiche sala
            stati = {}
            for t in tavoli_detail:
                s = t["status"]
                stati[s] = stati.get(s, 0) + 1

            # Fatturato sala oggi
            fatturato_sala = 0
            for t in tavoli:
                comande_chiuse = db.query(Comanda).filter(
                    Comanda.tavolo_id == t.id,
                    Comanda.status == "chiusa",
                    Comanda.closed_at >= today
                ).all()
                fatturato_sala += sum(c.totale for c in comande_chiuse)

            sale_data.append({
                "sala_id":   sala.id,
                "sala_nome": sala.nome,
                "outlet_id": sala.outlet_id,
                "tavoli":    tavoli_detail,
                "stats":     stati,
                "totale_tavoli": len(tavoli),
                "fatturato_oggi": round(fatturato_sala, 2),
            })

        # ── Andamento prenotazioni per ora ────────────────────────────────────
        pren_oggi = db.query(Prenotazione).filter(
            Prenotazione.data == today
        ).all()

        # Raggruppa per ora (08, 09, 10, ... 23)
        pren_per_ora = {}
        coperti_per_ora = {}
        for p in pren_oggi:
            try:
                ora = p.orario.split(":")[0] if p.orario else "00"
                ora = ora.zfill(2)
                pren_per_ora[ora]    = pren_per_ora.get(ora, 0) + 1
                coperti_per_ora[ora] = coperti_per_ora.get(ora, 0) + p.coperti
            except: pass

        # Slot con prenotazioni ordinati
        ore_attive = sorted(set(list(pren_per_ora.keys()) + list(coperti_per_ora.keys())))
        andamento = [
            {"ora": o, "prenotazioni": pren_per_ora.get(o,0), "coperti": coperti_per_ora.get(o,0)}
            for o in ore_attive
        ]

        # ── Prenotazioni per turno ────────────────────────────────────────────
        turni_oggi = db.query(Turno).all()
        turni_stats = []
        for turno in turni_oggi:
            q = db.query(Prenotazione).filter(
                Prenotazione.data == today,
                Prenotazione.turno_id == turno.id,
            )
            tot_pren   = q.count()
            tot_coperti = sum(p.coperti for p in q.all())
            cap         = turno.copertura_max or 0
            turni_stats.append({
                "turno_id":   turno.id,
                "nome":       turno.nome,
                "servizio":   turno.servizio,
                "ora_inizio": turno.ora_inizio,
                "ora_fine":   turno.ora_fine,
                "prenotazioni": tot_pren,
                "coperti":    tot_coperti,
                "capacita_max": cap,
                "perc_riempimento": round((tot_coperti/cap)*100) if cap>0 else None,
            })

        # ── Riepilogo globale ─────────────────────────────────────────────────
        tot_liberi    = sum(t["stats"].get("disponibile",0) for t in sale_data)
        tot_occupati  = sum(t["stats"].get("occupato",0)    for t in sale_data)
        tot_conto     = sum(t["stats"].get("chiesto_conto",0)+t["stats"].get("uscita",0) for t in sale_data)
        tot_riservati = sum(t["stats"].get("riservato",0)   for t in sale_data)
        tot_pren_oggi = len(pren_oggi)
        tot_coperti_pren = sum(p.coperti for p in pren_oggi)
        tot_vip       = sum(1 for p in pren_oggi if getattr(p,"is_vip",False))
        fat_totale    = sum(s["fatturato_oggi"] for s in sale_data)

        # ── Prossime prenotazioni ─────────────────────────────────────────────
        pren_prossime = sorted(
            [{"nome":p.nome,"ora":p.orario,"coperti":p.coperti,
              "stato":"confermata" if p.confermata else "in_attesa"}
             for p in pren_oggi if p.orario],
            key=lambda x: x["ora"]
        )

        return ok({
            "ora_aggiornamento": now.strftime("%H:%M:%S"),
            "sale":  sale_data,
            "andamento_prenotazioni": andamento,
            "turni_stats": turni_stats,
            "prenotazioni_prossime": pren_prossime,
            "riepilogo": {
                "tavoli_liberi":    tot_liberi,
                "tavoli_occupati":  tot_occupati,
                "tavoli_conto":     tot_conto,
                "tavoli_riservati": tot_riservati,
                "prenotazioni_oggi":tot_pren_oggi,
                "coperti_previsti": tot_coperti_pren,
                "ospiti_vip":       tot_vip,
                "fatturato_oggi":   round(fat_totale, 2),
            }
        })
    finally:
        db.close()
